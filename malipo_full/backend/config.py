"""
=============================================================================
KENYA PAYROLL SYSTEM — Central Configuration
=============================================================================
Single source of truth for:
  • Statutory constants (update here when Finance Act changes)
  • Directory layout
  • Authority portal metadata
  • Extension hooks for Phases 6 (Audit/Backup) and 7 (Security)

Design principle: every constant that could change with a Finance Act or
KRA circular is isolated here, not buried in calculation code.
=============================================================================
"""

from __future__ import annotations
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
import os

# ─── BASE PATHS ──────────────────────────────────────────────────────────────
ROOT_DIR      = Path(__file__).parent
SHARD_DIR     = ROOT_DIR / "shards"
TEMPLATE_DIR  = ROOT_DIR / "templates"      # Phase 4: downloaded authority templates
FILLED_DIR    = ROOT_DIR / "filled"         # Phase 4: populated templates
EXPORT_DIR    = ROOT_DIR / "exports"        # Phase 5: aggregated reports
BACKUP_DIR    = ROOT_DIR / "backups"        # Phase 6: shard snapshots
LOG_DIR       = ROOT_DIR / "logs"

for _d in (SHARD_DIR, TEMPLATE_DIR, FILLED_DIR, EXPORT_DIR, BACKUP_DIR, LOG_DIR):
    _d.mkdir(parents=True, exist_ok=True)


# ─── SHARDING ─────────────────────────────────────────────────────────────────
EMPLOYEES_PER_SHARD = 5_000


# ─── 2026 STATUTORY CONSTANTS ────────────────────────────────────────────────
# Update this section when Finance Act or KRA circulars change rates.

@dataclass(frozen=True)
class NSSFConfig:
    """NSSF Year 4 rates — effective Feb 1 2026."""
    tier1_ceiling:   float = 9_000      # KES
    tier2_ceiling:   float = 108_000    # KES (Upper Earnings Limit)
    rate:            float = 0.06       # 6% both tiers, both sides
    tier1_max:       float = 540.0
    tier2_max:       float = 5_940.0
    employee_max:    float = 6_480.0    # tier1_max + tier2_max

NSSF = NSSFConfig()


@dataclass(frozen=True)
class SHIFConfig:
    """Social Health Insurance Fund."""
    rate:    float = 0.0275    # 2.75% of gross (uncapped)
    minimum: float = 0.0       # No statutory minimum as of 2026

SHIF = SHIFConfig()


@dataclass(frozen=True)
class AHLConfig:
    """Affordable Housing Levy — both sides 1.5%."""
    employee_rate: float = 0.015
    employer_rate: float = 0.015
    relief_rate:   float = 0.15     # 15% of employee AHL applied as PAYE relief

AHL = AHLConfig()


@dataclass(frozen=True)
class PAYEConfig:
    """2026 progressive PAYE bands and standard reliefs."""
    # (upper_limit_KES, marginal_rate)
    bands: tuple = (
        (24_000,   0.10),
        (32_333,   0.25),
        (500_000,  0.30),
        (800_000,  0.325),
        (float("inf"), 0.35),
    )
    nonresident_rate:        float = 0.30

    personal_relief:         float = 2_400.0   # per month
    insurance_relief_cap:    float = 5_000.0   # per month
    mortgage_relief_cap:     float = 30_000.0  # per month (KES 360,000/yr) — CORRECTED
    post_retirement_cap:     float = 15_000.0  # per month — CORRECTED: direct deduction cap
    disability_exemption_cap:float = 150_000.0 # per month
    pension_pre_tax_cap:     float = 30_000.0  # per month (KES 360,000/yr) — NEW

PAYE = PAYEConfig()


@dataclass(frozen=True)
class FilingConfig:
    """Filing deadlines and penalties."""
    paye_deadline_day:   int   = 9      # 9th of following month
    nssf_deadline_day:   int   = 9
    shif_deadline_day:   int   = 9
    ahl_deadline_day:    int   = 9      # 9th working day
    shif_late_penalty:   float = 0.02   # 2% per month
    ahl_late_penalty:    float = 0.03   # 3% per month
    paye_late_penalty_pct: float = 0.05   # 5% of tax due (one-time)
    paye_late_interest:    float = 0.01   # 1% per month on unpaid balance
    paye_min_penalty:      float = 10_000.0 # minimum KES 10,000

FILING = FilingConfig()


# ─── AUTHORITY PORTAL METADATA ───────────────────────────────────────────────
# Used by Phase 4 (template downloader) and Phase 7 (secure credential store).

@dataclass
class PortalConfig:
    code:        str
    name:        str
    base_url:    str
    login_url:   str
    template_url: str
    paybill:     str
    color_hex:   str    # for report styling

PORTALS: dict[str, PortalConfig] = {
    "kra": PortalConfig(
        code         = "kra",
        name         = "KRA iTax",
        base_url     = "https://itax.kra.go.ke",
        login_url    = "https://itax.kra.go.ke/KRA-Portal/",
        template_url = "https://itax.kra.go.ke/KRA-Portal/payeReturn.htm",
        paybill      = "222222",  # Government PayBill (Presidential Directive, Kenya Gazette 16008)
        color_hex    = "#22c55e",
    ),
    "nssf": PortalConfig(
        code         = "nssf",
        name         = "NSSF Portal",
        base_url     = "https://www.nssf.or.ke",
        login_url    = "https://www.nssf.or.ke/login",
        template_url = "https://www.nssf.or.ke/employer/contributions",
        paybill      = "200777",
        color_hex    = "#38bdf8",
    ),
    "shif": PortalConfig(
        code         = "shif",
        name         = "SHA Portal",
        base_url     = "https://sha.go.ke",
        login_url    = "https://sha.go.ke/employer/login",
        template_url = "https://sha.go.ke/employer/contributions/template",
        paybill      = "363636",
        color_hex    = "#f59e0b",
    ),
    "ahl": PortalConfig(
        code         = "ahl",
        name         = "KRA iTax (AHL)",
        base_url     = "https://itax.kra.go.ke",
        login_url    = "https://itax.kra.go.ke/KRA-Portal/",
        template_url = "https://itax.kra.go.ke/KRA-Portal/ahlReturn.htm",
        paybill      = "222222",  # Government PayBill (Presidential Directive, Kenya Gazette 16008)
        color_hex    = "#a78bfa",
    ),
}


# ─── WORKER QUEUE TUNING ─────────────────────────────────────────────────────
WORKER_MAX_RETRIES      = 3
WORKER_RETRY_BASE_DELAY = 0.25   # seconds; doubles each retry
WORKER_QUEUE_TIMEOUT    = 2.0    # seconds before checking stop flag


# ─── PHASE 5: AGGREGATOR ─────────────────────────────────────────────────────
# Tolerance for floating-point reconciliation across shards
AGGREGATOR_TOLERANCE_KES = 0.05   # 5 cents max rounding drift per shard

# GL account codes — override per company in Settings (Phase 7)
@dataclass
class GLAccounts:
    gross_payroll:     str = "5000"
    paye_liability:    str = "2100"
    nssf_employee:     str = "2110"
    nssf_employer:     str = "5110"
    shif_liability:    str = "2120"
    ahl_employee:      str = "2130"
    ahl_employer:      str = "5130"
    net_pay_liability: str = "2200"
    bank_clearing:     str = "1050"

GL = GLAccounts()


# ─── PHASE 6: AUDIT & BACKUP HOOKS ───────────────────────────────────────────
# These are wired up in Phase 6 but declared here so Phases 4/5 can reference.
BACKUP_RETENTION_DAYS   = 90
BACKUP_SCHEDULE_CRON    = "0 2 * * *"    # 02:00 daily
MAX_BACKUP_SIZE_MB      = 500


# ─── PHASE 7: SECURITY HOOKS ─────────────────────────────────────────────────
# Credential store interface — implemented in Phase 7 but referenced here.
# Phases 4/5 call get_portal_credential() which returns from env by default.

def get_portal_credential(portal_code: str, field: str) -> Optional[str]:
    """
    Retrieve a portal credential.
    Phase 7 will replace this with an encrypted vault lookup.
    Falls back to environment variables: PAYROLL_{PORTAL}_{FIELD}
    e.g. PAYROLL_KRA_USERNAME, PAYROLL_KRA_PASSWORD
    """
    env_key = f"PAYROLL_{portal_code.upper()}_{field.upper()}"
    return os.environ.get(env_key)


# ─── EXCEL STYLING CONSTANTS ─────────────────────────────────────────────────
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

STYLE = {
    "header_fill":   PatternFill("solid", fgColor="1a2e20"),
    "header_font":   Font(bold=True, color="22c55e", name="Calibri", size=10),
    "title_font":    Font(bold=True, color="FFFFFF", name="Calibri", size=12),
    "total_fill":    PatternFill("solid", fgColor="0b1610"),
    "total_font":    Font(bold=True, color="f59e0b", name="Calibri", size=10),
    "data_font":     Font(name="Calibri", size=10, color="dff0e6"),
    "data_fill":     PatternFill("solid", fgColor="101f15"),
    "alt_fill":      PatternFill("solid", fgColor="0d1a10"),
    "border_side":   Side(style="thin", color="1a2e20"),
    "money_fmt":     '#,##0.00',
    "int_fmt":       '#,##0',
    "pct_fmt":       '0.00%',
    "date_fmt":      'DD-MMM-YYYY',
    "align_right":   Alignment(horizontal="right", vertical="center"),
    "align_center":  Alignment(horizontal="center", vertical="center"),
    "align_left":    Alignment(horizontal="left", vertical="center"),
}

def make_border():
    s = STYLE["border_side"]
    return Border(left=s, right=s, top=s, bottom=s)
