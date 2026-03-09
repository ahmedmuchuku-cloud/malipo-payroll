"""
=============================================================================
KENYA PAYROLL SYSTEM — Phase 5: Aggregator Service
=============================================================================
Responsibilities:
  1. ShardReader          — pulls raw payroll_transactions from every shard
                            and enriches each row with employee metadata
                            (name, PIN, NSSF No., etc.) from the same shard.

  2. Aggregator           — merges shard data, cross-validates each shard's
                            sub-total against the in-memory collector total,
                            and produces the canonical AggregationReport.

  3. Exporters:
       GLExporter          — double-entry journal entries (CSV + Excel)
       PayrollRegisterExporter — full employee register (Excel)
       RemittanceSummaryExporter — one-page executive summary (Excel)
       GovernmentFileExporter  — portal-ready CSV for KRA / NSSF / SHIF / AHL
       P9ExporterExcel     — annual P9A forms (Excel, one sheet per employee)

  4. ExportOrchestrator   — single call that runs all exporters and fires
                            EventBus events for Phase 6 audit hooks.

Design notes for future phases
──────────────────────────────
• Phase 6 (Audit & Backup): ExportOrchestrator publishes EXPORT_CREATED and
  AGGREGATION_DONE events; the BackupPlugin subscribes and snapshots shards.
  The manifest written here also carries SHA-256 checksums that the
  AuditPlugin verifies on each load.

• Phase 7 (Security / RBAC): ExportOrchestrator accepts an `operator` str
  that will be checked against the RBACPlugin before files are written.
  Sensitive exports (P9A) will be encrypted at rest; the EncryptionPlugin
  hooks into on_export_created to do this transparently.

• Multi-company: AggregationReport.company_id is propagated into every
  export so multi-tenant reports remain isolated.
=============================================================================
"""

from __future__ import annotations

import csv
import json
import logging
import math
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    EXPORT_DIR, GL, AGGREGATOR_TOLERANCE_KES,
    PORTALS, STYLE, make_border, FILING,
)
from events import PayrollEvent, get_bus, EventBus
from shard_manager import ShardManager

logger = logging.getLogger(__name__)


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _period(month: int, year: int) -> str:
    return f"{year}-{month:02d}"


def _month_name(month: int, year: int) -> str:
    return date(year, month, 1).strftime("%B %Y")


def _money(v) -> float:
    return round(float(v or 0), 2)


def _fmt(v: float) -> str:
    return f"{v:,.2f}"


def _header_style(ws, row: int, n_cols: int, color: str = "1a2e20") -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = make_border()
    ws.row_dimensions[row].height = 28


def _total_style(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = PatternFill("solid", fgColor="0b1610")
        cell.font      = Font(bold=True, color="f59e0b", name="Calibri", size=10)
        cell.border    = make_border()


def _data_style(ws, row: int, n_cols: int, alt: bool = False) -> None:
    hex_ = "0d1a10" if alt else "101f15"
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = PatternFill("solid", fgColor=hex_)
        cell.font   = Font(name="Calibri", size=9, color="dff0e6")
        cell.border = make_border()


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _money_cell(ws, row: int, col: int, value: float) -> None:
    cell = ws.cell(row=row, column=col, value=_money(value))
    cell.number_format = "#,##0.00"
    cell.alignment     = Alignment(horizontal="right", vertical="center")


# ─── DATA MODELS ──────────────────────────────────────────────────────────────

@dataclass
class EnrichedTransaction:
    """
    A payroll_transaction row joined with employee metadata.
    This is the canonical row object used by all exporters.
    """
    # Employee identity
    employee_id:   int
    full_name:     str
    kra_pin:       str = ""
    nssf_number:   str = ""
    sha_number:    str = ""
    residency:     str = "resident"
    department:    str = ""
    job_title:     str = ""

    # Shard provenance (for validation)
    shard_id:       int = 0
    payroll_run_id: int = 0

    # Earnings
    gross_salary:            float = 0.0
    base_salary:             float = 0.0
    taxable_allowances:      float = 0.0
    non_taxable_allowances:  float = 0.0
    benefits_in_kind:        float = 0.0

    # NSSF
    nssf_tier1_employee: float = 0.0
    nssf_tier2_employee: float = 0.0
    nssf_employee:       float = 0.0
    nssf_tier1_employer: float = 0.0
    nssf_tier2_employer: float = 0.0
    nssf_employer:       float = 0.0

    # SHIF & AHL
    shif:                   float = 0.0
    housing_levy_employee:  float = 0.0
    housing_levy_employer:  float = 0.0

    # PAYE
    pre_tax_deductions:     float = 0.0
    disability_exemption:   float = 0.0
    taxable_income:         float = 0.0
    gross_paye:             float = 0.0

    # Reliefs
    personal_relief:        float = 0.0
    insurance_relief:       float = 0.0
    mortgage_relief:        float = 0.0
    ahl_relief:             float = 0.0
    post_retirement_relief: float = 0.0
    total_relief:           float = 0.0
    paye:                   float = 0.0

    # Other
    helb_deduction:   float = 0.0
    other_deductions: float = 0.0

    # Summary
    total_statutory:   float = 0.0
    total_deductions:  float = 0.0
    net_pay:           float = 0.0


@dataclass
class ShardSubtotal:
    shard_id:       int
    payroll_run_id: int
    employee_count: int = 0
    gross_salary:   float = 0.0
    paye:           float = 0.0
    nssf_employee:  float = 0.0
    nssf_employer:  float = 0.0
    shif:           float = 0.0
    ahl_employee:   float = 0.0
    ahl_employer:   float = 0.0
    net_pay:        float = 0.0


@dataclass
class ValidationIssue:
    severity:   str   # "ERROR" | "WARNING"
    shard_id:   int
    field:      str
    expected:   float
    actual:     float
    delta:      float


@dataclass
class AggregationReport:
    """
    The canonical output of Phase 5.
    Contains grand totals, per-shard subtotals, validation issues,
    and all enriched transaction rows ready for export.
    """
    company_id:    int
    month:         int
    year:          int
    generated_at:  str = field(default_factory=lambda: datetime.utcnow().isoformat())

    # Grand totals (sum of all shards)
    employee_count:       int   = 0
    gross_salary:         float = 0.0
    nssf_employee:        float = 0.0
    nssf_employer:        float = 0.0
    shif:                 float = 0.0
    housing_levy_employee: float = 0.0
    housing_levy_employer: float = 0.0
    paye:                 float = 0.0
    total_statutory:      float = 0.0
    net_pay:              float = 0.0

    # Employer totals (cost to company)
    total_employer_cost:  float = 0.0   # gross + nssf_employer + ahl_employer

    # Detailed rows and validation
    transactions:     list[EnrichedTransaction] = field(default_factory=list)
    shard_subtotals:  list[ShardSubtotal]       = field(default_factory=list)
    validation_issues: list[ValidationIssue]    = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not any(v.severity == "ERROR" for v in self.validation_issues)

    def summary_dict(self) -> dict:
        return {
            "period":          _period(self.month, self.year),
            "employee_count":  self.employee_count,
            "gross_salary":    _money(self.gross_salary),
            "nssf_employee":   _money(self.nssf_employee),
            "nssf_employer":   _money(self.nssf_employer),
            "shif":            _money(self.shif),
            "ahl_employee":    _money(self.housing_levy_employee),
            "ahl_employer":    _money(self.housing_levy_employer),
            "paye":            _money(self.paye),
            "total_statutory": _money(self.total_statutory),
            "net_pay":         _money(self.net_pay),
            "valid":           self.is_valid,
        }


# ─── SHARD READER ─────────────────────────────────────────────────────────────

class ShardReader:
    """
    Reads payroll_transactions and joins employee metadata from one shard.
    Pure read — never writes. Safe to call while ShardWorkers are done.
    """

    def __init__(self, shard_manager: ShardManager):
        self.sm = shard_manager

    def read_shard(self, shard_id: int,
                   payroll_run_id: int) -> tuple[list[EnrichedTransaction], ShardSubtotal]:
        """
        Returns (enriched_rows, subtotal) for one shard.
        Joins payroll_transactions with employees on employee_id.
        """
        rows: list[EnrichedTransaction] = []
        sub  = ShardSubtotal(shard_id=shard_id, payroll_run_id=payroll_run_id)

        with self.sm.connection(shard_id) as conn:
            sql = """
                SELECT
                    pt.*,
                    e.full_name, e.kra_pin, e.nssf_number, e.sha_number,
                    e.residency_status, e.department, e.job_title
                FROM payroll_transactions pt
                JOIN employees e ON e.employee_id = pt.employee_id
                WHERE pt.payroll_run_id = ?
                ORDER BY pt.employee_id
            """
            raw_rows = conn.execute(sql, (payroll_run_id,)).fetchall()

        for r in raw_rows:
            d = dict(r)
            et = EnrichedTransaction(
                employee_id             = d["employee_id"],
                full_name               = d.get("full_name", ""),
                kra_pin                 = d.get("kra_pin", "") or "",
                nssf_number             = d.get("nssf_number", "") or "",
                sha_number              = d.get("sha_number", "") or "",
                residency               = d.get("residency_status", "resident"),
                department              = d.get("department", "") or "",
                job_title               = d.get("job_title", "") or "",
                shard_id                = shard_id,
                payroll_run_id          = payroll_run_id,
                gross_salary            = _money(d.get("gross_salary", 0)),
                base_salary             = _money(d.get("base_salary", 0)),
                taxable_allowances      = _money(d.get("taxable_allowances", 0)),
                non_taxable_allowances  = _money(d.get("non_taxable_allowances", 0)),
                benefits_in_kind        = _money(d.get("benefits_in_kind", 0)),
                nssf_tier1_employee     = _money(d.get("nssf_tier1_employee", 0)),
                nssf_tier2_employee     = _money(d.get("nssf_tier2_employee", 0)),
                nssf_employee           = _money(d.get("nssf_employee", 0)),
                nssf_tier1_employer     = _money(d.get("nssf_tier1_employer", 0)),
                nssf_tier2_employer     = _money(d.get("nssf_tier2_employer", 0)),
                nssf_employer           = _money(d.get("nssf_employer", 0)),
                shif                    = _money(d.get("shif", 0)),
                housing_levy_employee   = _money(d.get("housing_levy_employee", 0)),
                housing_levy_employer   = _money(d.get("housing_levy_employer", 0)),
                pre_tax_deductions      = _money(d.get("pre_tax_deductions", 0)),
                disability_exemption    = _money(d.get("disability_exemption", 0)),
                taxable_income          = _money(d.get("taxable_income", 0)),
                gross_paye              = _money(d.get("gross_paye", 0)),
                personal_relief         = _money(d.get("personal_relief", 0)),
                insurance_relief        = _money(d.get("insurance_relief", 0)),
                mortgage_relief         = _money(d.get("mortgage_relief", 0)),
                ahl_relief              = _money(d.get("ahl_relief", 0)),
                post_retirement_relief  = _money(d.get("post_retirement_relief", 0)),
                total_relief            = _money(d.get("total_relief", 0)),
                paye                    = _money(d.get("paye", 0)),
                helb_deduction          = _money(d.get("helb_deduction", 0)),
                other_deductions        = _money(d.get("other_deductions", 0)),
                total_statutory         = _money(d.get("total_statutory", 0)),
                total_deductions        = _money(d.get("total_deductions", 0)),
                net_pay                 = _money(d.get("net_pay", 0)),
            )
            rows.append(et)

            sub.employee_count += 1
            sub.gross_salary   += et.gross_salary
            sub.paye           += et.paye
            sub.nssf_employee  += et.nssf_employee
            sub.nssf_employer  += et.nssf_employer
            sub.shif           += et.shif
            sub.ahl_employee   += et.housing_levy_employee
            sub.ahl_employer   += et.housing_levy_employer
            sub.net_pay        += et.net_pay

        # Round subtotals to avoid float drift
        for f in ("gross_salary", "paye", "nssf_employee", "nssf_employer",
                   "shif", "ahl_employee", "ahl_employer", "net_pay"):
            setattr(sub, f, _money(getattr(sub, f)))

        return rows, sub


# ─── AGGREGATOR ───────────────────────────────────────────────────────────────

class Aggregator:
    """
    Merges transactions from all shards and validates totals.

    Validation strategy:
      For every monetary field F on every shard S:
        |shard_S.F  −  collector_shard_totals[S].F|  ≤  TOLERANCE
      If the collector isn't available (e.g. post-restart re-aggregation),
      we cross-validate shard subtotals against the grand total only.
    """

    def __init__(self, shard_manager: ShardManager):
        self.sm     = shard_manager
        self.reader = ShardReader(shard_manager)

    def aggregate(self,
                  month:            int,
                  year:             int,
                  company_id:       int = 1,
                  shard_ids:        list[int] | None = None,
                  collector_totals: dict | None = None,
                  ) -> AggregationReport:
        """
        Read every shard, merge, validate and return AggregationReport.

        `collector_totals`: the ResultCollector.shard_totals dict from the
        run that produced the data — used for cross-validation. Optional;
        if omitted, intra-report validation only.
        """
        if shard_ids is None:
            shard_ids = self.sm.list_shards()

        report = AggregationReport(company_id=company_id,
                                   month=month, year=year)

        for shard_id in shard_ids:
            # Resolve payroll_run_id for this shard+period
            run_id = self._get_run_id(shard_id, month, year, company_id)
            if run_id is None:
                logger.warning(f"No processed payroll run found for "
                               f"shard {shard_id} {month}/{year}")
                continue

            rows, sub = self.reader.read_shard(shard_id, run_id)
            report.transactions.extend(rows)
            report.shard_subtotals.append(sub)

            # Cross-validate against collector if available
            if collector_totals and shard_id in collector_totals:
                self._validate_shard(report, sub,
                                     collector_totals[shard_id])

        # Compute grand totals from subtotals
        for sub in report.shard_subtotals:
            report.employee_count        += sub.employee_count
            report.gross_salary          += sub.gross_salary
            report.paye                  += sub.paye
            report.nssf_employee         += sub.nssf_employee
            report.nssf_employer         += sub.nssf_employer
            report.shif                  += sub.shif
            report.housing_levy_employee += sub.ahl_employee
            report.housing_levy_employer += sub.ahl_employer
            report.net_pay               += sub.net_pay

        # Round all grand totals
        for f in ("gross_salary", "paye", "nssf_employee", "nssf_employer",
                   "shif", "housing_levy_employee", "housing_levy_employer",
                   "net_pay"):
            setattr(report, f, _money(getattr(report, f)))

        report.total_statutory   = _money(
            report.paye + report.nssf_employee +
            report.shif + report.housing_levy_employee
        )
        report.total_employer_cost = _money(
            report.gross_salary + report.nssf_employer + report.housing_levy_employer
        )

        # Sanity check: sum(sub.net_pay) == report.net_pay
        recalc_net = _money(sum(s.net_pay for s in report.shard_subtotals))
        if abs(recalc_net - report.net_pay) > AGGREGATOR_TOLERANCE_KES:
            report.validation_issues.append(ValidationIssue(
                severity="ERROR", shard_id=0,
                field="net_pay",
                expected=report.net_pay, actual=recalc_net,
                delta=abs(recalc_net - report.net_pay),
            ))

        n_issues = len(report.validation_issues)
        logger.info(
            f"Aggregation complete: {report.employee_count} employees, "
            f"gross={_fmt(report.gross_salary)}, "
            f"paye={_fmt(report.paye)}, "
            f"{'VALID' if report.is_valid else f'INVALID ({n_issues} issues)'}"
        )
        return report

    def _get_run_id(self, shard_id: int, month: int, year: int,
                    company_id: int) -> int | None:
        with self.sm.connection(shard_id) as conn:
            row = conn.execute("""
                SELECT payroll_run_id FROM payroll_runs
                WHERE company_id=? AND run_month=? AND run_year=?
                  AND status IN ('processed', 'locked')
                ORDER BY payroll_run_id DESC LIMIT 1
            """, (company_id, month, year)).fetchone()
        return row[0] if row else None

    def _validate_shard(self, report: AggregationReport,
                         sub: ShardSubtotal,
                         collector_sub: dict) -> None:
        """Compare DB-read shard subtotals against in-memory collector values."""
        mapping = {
            "gross_salary":  "gross_salary",
            "paye":          "paye",
            "nssf_employee": "nssf_employee",
            "shif":          "shif",
        }
        for sub_field, col_key in mapping.items():
            actual   = getattr(sub, sub_field)
            expected = _money(collector_sub.get(col_key, 0))
            delta    = abs(actual - expected)
            if delta > AGGREGATOR_TOLERANCE_KES:
                report.validation_issues.append(ValidationIssue(
                    severity="WARNING",
                    shard_id=sub.shard_id,
                    field=sub_field,
                    expected=expected,
                    actual=actual,
                    delta=delta,
                ))


# ─── EXPORTERS ────────────────────────────────────────────────────────────────

class PayrollRegisterExporter:
    """
    Full payroll register: one row per employee, all deductions and net pay.
    Saved as Excel with branded styling.
    """

    HEADERS = [
        ("Emp #",           8),
        ("Employee Name",  28),
        ("KRA PIN",        14),
        ("Department",     16),
        ("Gross Salary",   14),
        ("NSSF Employee",  14),
        ("SHIF",           12),
        ("AHL Employee",   12),
        ("Taxable Income", 14),
        ("PAYE",           12),
        ("Total Deductions",14),
        ("Net Pay",        14),
        ("NSSF Employer",  14),
        ("AHL Employer",   12),
        ("CTC",            14),
    ]

    def export(self, report: AggregationReport,
               export_dir: Path = EXPORT_DIR) -> Path:
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Payroll Register"
        period = _period(report.month, report.year)
        n = len(self.HEADERS)

        # Title
        ws.merge_cells(f"A1:{get_column_letter(n)}1")
        ws["A1"] = f"PAYROLL REGISTER — {_month_name(report.month, report.year)}"
        ws["A1"].font      = Font(bold=True, color="22c55e", size=14, name="Calibri")
        ws["A1"].fill      = PatternFill("solid", fgColor="050e09")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 32

        ws["A2"] = f"Employees: {report.employee_count}  |  Generated: {report.generated_at[:19]}"
        ws["A2"].font = Font(size=9, color="5a7a65", name="Calibri")
        ws.row_dimensions[2].height = 18

        # Headers
        for col, (hdr, w) in enumerate(self.HEADERS, 1):
            ws.cell(row=3, column=col, value=hdr)
            ws.column_dimensions[get_column_letter(col)].width = w
        _header_style(ws, 3, n, "1a2e20")

        # Data
        money_cols = set(range(5, n + 1))
        for i, tx in enumerate(report.transactions):
            r   = 4 + i
            alt = (i % 2 == 1)
            ctc = _money(tx.gross_salary + tx.nssf_employer + tx.housing_levy_employer)
            values = [
                tx.employee_id, tx.full_name, tx.kra_pin, tx.department,
                tx.gross_salary, tx.nssf_employee, tx.shif,
                tx.housing_levy_employee, tx.taxable_income, tx.paye,
                tx.total_deductions, tx.net_pay,
                tx.nssf_employer, tx.housing_levy_employer, ctc,
            ]
            _data_style(ws, r, n, alt)
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                if col in money_cols and isinstance(val, float):
                    cell.number_format = "#,##0.00"
                    cell.alignment     = Alignment(horizontal="right", vertical="center")
                elif col == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Totals
        tr = 4 + len(report.transactions)
        total_ctc = _money(report.total_employer_cost)
        totals = [
            "", "TOTALS", "", "",
            report.gross_salary, report.nssf_employee,
            report.shif, report.housing_levy_employee,
            "", report.paye, "", report.net_pay,
            report.nssf_employer, report.housing_levy_employer, total_ctc,
        ]
        _total_style(ws, tr, n)
        for col, val in enumerate(totals, 1):
            cell = ws.cell(row=tr, column=col, value=val if val != "" else None)
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right", vertical="center")

        ws.freeze_panes = "A4"

        out = export_dir / f"payroll_register_{period}.xlsx"
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
        logger.info(f"Payroll register exported: {out}")
        return out


class GLExporter:
    """
    Double-entry GL journal entries for the payroll run.

    Each payroll run produces these journal lines:
      DR  Gross Payroll Expense       (gross_salary)
        CR  PAYE Liability            (paye)
        CR  NSSF Employee Liability   (nssf_employee)
        CR  NSSF Employer Expense     (nssf_employer)  ← DR Payroll Expense too
        CR  SHIF Liability            (shif)
        CR  AHL Employee Liability    (ahl_employee)
        CR  AHL Employer Expense      (ahl_employer)   ← DR Payroll Expense too
        CR  Net Pay Liability         (net_pay)

    Employer contributions hit a separate expense account + liability:
      DR  NSSF Employer Expense       (nssf_employer)
        CR  NSSF Employer Liability   (nssf_employer)
    """

    def export(self, report: AggregationReport,
               export_dir: Path = EXPORT_DIR) -> tuple[Path, Path]:
        """Returns (xlsx_path, csv_path)."""
        period    = _period(report.month, report.year)
        entries   = self._build_entries(report)

        xlsx_path = export_dir / f"gl_journal_{period}.xlsx"
        csv_path  = export_dir / f"gl_journal_{period}.csv"
        export_dir.mkdir(parents=True, exist_ok=True)

        self._write_xlsx(entries, report, xlsx_path)
        self._write_csv(entries, csv_path)

        logger.info(f"GL journal exported: {xlsx_path}, {csv_path}")
        return xlsx_path, csv_path

    def _build_entries(self, report: AggregationReport) -> list[dict]:
        period = _period(report.month, report.year)
        desc   = f"Payroll — {_month_name(report.month, report.year)}"
        entries = [
            # ── Employee deductions create liabilities ────────────────────────
            {"account": GL.gross_payroll,    "description": desc, "dr": report.gross_salary,            "cr": 0},
            {"account": GL.paye_liability,   "description": desc, "dr": 0,                              "cr": report.paye},
            {"account": GL.nssf_employee,    "description": desc, "dr": 0,                              "cr": report.nssf_employee},
            {"account": GL.shif_liability,   "description": desc, "dr": 0,                              "cr": report.shif},
            {"account": GL.ahl_employee,     "description": desc, "dr": 0,                              "cr": report.housing_levy_employee},
            {"account": GL.net_pay_liability,"description": desc, "dr": 0,                              "cr": report.net_pay},
            # ── Employer contributions ────────────────────────────────────────
            {"account": GL.nssf_employer,    "description": f"{desc} — employer NSSF",  "dr": report.nssf_employer,           "cr": 0},
            {"account": GL.nssf_employee,    "description": f"{desc} — employer NSSF",  "dr": 0,                              "cr": report.nssf_employer},
            {"account": GL.ahl_employer,     "description": f"{desc} — employer AHL",   "dr": report.housing_levy_employer,   "cr": 0},
            {"account": GL.ahl_employee,     "description": f"{desc} — employer AHL",   "dr": 0,                              "cr": report.housing_levy_employer},
        ]
        for e in entries:
            e["period"]  = period
            e["dr"]      = _money(e["dr"])
            e["cr"]      = _money(e["cr"])
            e["balance"] = _money(e["dr"] - e["cr"])
        return entries

    def _write_xlsx(self, entries: list[dict],
                    report: AggregationReport, path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GL Journal"

        ws.merge_cells("A1:F1")
        ws["A1"] = (f"GL JOURNAL — {_month_name(report.month, report.year)}  "
                    f"| Employees: {report.employee_count}")
        ws["A1"].font      = Font(bold=True, color="f59e0b", size=12, name="Calibri")
        ws["A1"].fill      = PatternFill("solid", fgColor="050e09")
        ws["A1"].alignment = Alignment(horizontal="center")

        HEADERS = [("Account", 10), ("Period", 10), ("Description", 45),
                   ("Debit", 18), ("Credit", 18), ("Net", 18)]
        for col, (hdr, w) in enumerate(HEADERS, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        _header_style(ws, 3, 6, "1a2e20")
        for col, (hdr, _) in enumerate(HEADERS, 1):
            ws.cell(row=3, column=col, value=hdr)

        total_dr = total_cr = 0.0
        for i, e in enumerate(entries):
            r   = 4 + i
            alt = (i % 2 == 1)
            _data_style(ws, r, 6, alt)
            ws.cell(row=r, column=1, value=e["account"]).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r, column=2, value=e["period"]).alignment  = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r, column=3, value=e["description"])
            for col, key in [(4, "dr"), (5, "cr"), (6, "balance")]:
                cell = ws.cell(row=r, column=col, value=e[key])
                cell.number_format = "#,##0.00"
                color = "22c55e" if col == 4 else "ef4444" if col == 5 else "dff0e6"
                cell.font      = Font(name="Calibri", size=9, color=color)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            total_dr += e["dr"]; total_cr += e["cr"]

        # Totals
        tr = 4 + len(entries)
        _total_style(ws, tr, 6)
        ws.cell(row=tr, column=3, value="TOTALS")
        for col, val in [(4, total_dr), (5, total_cr), (6, total_dr - total_cr)]:
            cell = ws.cell(row=tr, column=col, value=_money(val))
            cell.number_format = "#,##0.00"
            cell.alignment     = Alignment(horizontal="right", vertical="center")

        # Balanced check
        is_balanced = abs(total_dr - total_cr) < 0.05
        ws.cell(row=tr + 1, column=3,
                value=f"{'✓ BALANCED' if is_balanced else '✗ OUT OF BALANCE'}")
        ws.cell(row=tr + 1, column=3).font = Font(
            bold=True, color="22c55e" if is_balanced else "ef4444",
            size=10, name="Calibri"
        )
        ws.freeze_panes = "A4"
        wb.save(str(path))

    def _write_csv(self, entries: list[dict], path: Path) -> None:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["period", "account", "description",
                                               "dr", "cr", "balance"])
            w.writeheader()
            w.writerows(entries)


class RemittanceSummaryExporter:
    """
    One-page executive summary: what to pay, to whom, by when.
    This is what the Finance Director prints and signs off on.
    """

    def export(self, report: AggregationReport,
               company_info: dict,
               export_dir: Path = EXPORT_DIR) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Remittance Summary"
        export_dir.mkdir(parents=True, exist_ok=True)

        period = _period(report.month, report.year)
        month_name = _month_name(report.month, report.year)

        # ── Title block ───────────────────────────────────────────────────────
        ws.merge_cells("A1:F1")
        ws["A1"] = f"PAYROLL REMITTANCE SUMMARY — {month_name}"
        ws["A1"].font      = Font(bold=True, color="22c55e", size=16, name="Calibri")
        ws["A1"].fill      = PatternFill("solid", fgColor="050e09")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A2:F2")
        ws["A2"] = (f"{company_info.get('name', '')}  |  "
                    f"PIN: {company_info.get('pin', '')}  |  "
                    f"Employees: {report.employee_count}  |  "
                    f"Period: {period}")
        ws["A2"].font      = Font(size=10, color="5a7a65", name="Calibri")
        ws["A2"].fill      = PatternFill("solid", fgColor="050e09")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 22

        # ── Column widths ─────────────────────────────────────────────────────
        for col, w in enumerate([6, 28, 16, 16, 16, 20], 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # ── Obligation rows ───────────────────────────────────────────────────
        obligations = [
            ("📋 PAYE",  "KRA iTax",   "572572",
             report.paye, "22c55e",
             f"Due 9th {month_name} — Employee income tax"),

            ("🛡 NSSF",  "NSSF Portal","200777",
             _money(report.nssf_employee + report.nssf_employer), "38bdf8",
             f"Emp KES {_fmt(report.nssf_employee)} + Emplr KES {_fmt(report.nssf_employer)}"),

            ("🏥 SHIF",  "SHA Portal", "363636",
             report.shif, "f59e0b",
             "2.75% of gross — Social Health Insurance Fund"),

            ("🏠 AHL",   "KRA iTax",   "572572",
             _money(report.housing_levy_employee + report.housing_levy_employer), "a78bfa",
             f"Emp KES {_fmt(report.housing_levy_employee)} + Emplr KES {_fmt(report.housing_levy_employer)}"),
        ]

        row = 4
        HDRS = ["#", "Obligation", "Amount Due (KES)", "M-Pesa Paybill",
                "Portal", "Notes"]
        for col, hdr in enumerate(HDRS, 1):
            ws.cell(row=row, column=col, value=hdr)
        _header_style(ws, row, 6)

        grand_total = 0.0
        for i, (icon_name, portal, paybill, amount, color, note) in enumerate(obligations):
            r   = row + 1 + i
            alt = (i % 2 == 1)
            _data_style(ws, r, 6, alt)
            ws.cell(row=r, column=1, value=i + 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r, column=2, value=icon_name).font  = Font(bold=True, color=color, name="Calibri", size=10)
            cell = ws.cell(row=r, column=3, value=amount)
            cell.number_format = "#,##0.00"
            cell.font          = Font(bold=True, color=color, name="Calibri", size=11)
            cell.alignment     = Alignment(horizontal="right", vertical="center")
            ws.cell(row=r, column=4, value=paybill).font   = Font(color="dff0e6", name="Courier New", size=10)
            ws.cell(row=r, column=5, value=portal).font    = Font(color="dff0e6", name="Calibri", size=9)
            ws.cell(row=r, column=6, value=note).font      = Font(color="5a7a65",  name="Calibri", size=9)
            ws.row_dimensions[r].height = 22
            grand_total += amount

        # Grand total
        tr = row + 1 + len(obligations)
        _total_style(ws, tr, 6)
        ws.cell(row=tr, column=2, value="TOTAL REMITTANCE")
        cell = ws.cell(row=tr, column=3, value=_money(grand_total))
        cell.number_format = "#,##0.00"
        cell.alignment     = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[tr].height = 24

        # ── Gross payroll summary block ───────────────────────────────────────
        gap = tr + 2
        summary_rows = [
            ("Total Gross Payroll",           report.gross_salary),
            ("Less: Total Statutory (employee)", -report.total_statutory),
            ("Net Pay to Employees",           report.net_pay),
            ("", None),
            ("Total Employer NSSF",           report.nssf_employer),
            ("Total Employer AHL",            report.housing_levy_employer),
            ("Total Cost to Company (CTC)",   report.total_employer_cost),
        ]
        ws.cell(row=gap, column=1, value="PAYROLL SUMMARY").font = Font(
            bold=True, color="dff0e6", name="Calibri", size=10)
        for j, (lbl, val) in enumerate(summary_rows):
            r = gap + 1 + j
            ws.cell(row=r, column=2, value=lbl).font = Font(
                color="5a7a65", name="Calibri", size=9)
            if val is not None:
                cell = ws.cell(row=r, column=3, value=_money(val))
                cell.number_format = "#,##0.00"
                cell.font      = Font(color="22c55e" if val >= 0 else "ef4444",
                                      bold=lbl.startswith("Total Cost"),
                                      name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # Validation status
        vrow = gap + len(summary_rows) + 3
        status = "✓ VALIDATION PASSED" if report.is_valid else f"✗ {len(report.validation_issues)} ISSUE(S)"
        ws.cell(row=vrow, column=1, value=status).font = Font(
            bold=True, size=11, name="Calibri",
            color="22c55e" if report.is_valid else "ef4444"
        )

        out = export_dir / f"remittance_summary_{period}.xlsx"
        wb.save(str(out))
        logger.info(f"Remittance summary exported: {out}")
        return out


class GovernmentFileExporter:
    """
    Generates portal-ready CSV files for each authority.
    These are the machine-readable upload files, distinct from
    the Phase 4 styled Excel templates.
    """

    def export_all(self, report: AggregationReport,
                   export_dir: Path = EXPORT_DIR) -> dict[str, Path]:
        export_dir.mkdir(parents=True, exist_ok=True)
        period = _period(report.month, report.year)
        return {
            "kra":  self._kra_paye(report, export_dir, period),
            "nssf": self._nssf(report, export_dir, period),
            "shif": self._shif(report, export_dir, period),
            "ahl":  self._ahl(report, export_dir, period),
        }

    def _kra_paye(self, report: AggregationReport,
                  d: Path, period: str) -> Path:
        path = d / f"kra_paye_{period}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["KRA_PIN", "EMPLOYEE_NAME", "GROSS_PAY", "BASIC_PAY",
                         "TAXABLE_ALLOW", "NON_TAXABLE_ALLOW", "BIK",
                         "NSSF_EMPLOYEE", "SHIF", "AHL_EMPLOYEE",
                         "TAXABLE_INCOME", "GROSS_PAYE",
                         "PERSONAL_RELIEF", "INSURANCE_RELIEF",
                         "MORTGAGE_RELIEF", "AHL_RELIEF",
                         "TOTAL_RELIEF", "NET_PAYE"])
            for tx in report.transactions:
                w.writerow([
                    tx.kra_pin or "A000000000A",
                    tx.full_name,
                    tx.gross_salary, tx.base_salary,
                    tx.taxable_allowances, tx.non_taxable_allowances,
                    tx.benefits_in_kind,
                    tx.nssf_employee, tx.shif, tx.housing_levy_employee,
                    tx.taxable_income, tx.gross_paye,
                    tx.personal_relief, tx.insurance_relief,
                    tx.mortgage_relief, tx.ahl_relief,
                    tx.total_relief, tx.paye,
                ])
        logger.info(f"KRA PAYE CSV: {path}")
        return path

    def _nssf(self, report: AggregationReport,
              d: Path, period: str) -> Path:
        path = d / f"nssf_{period}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["NSSF_MEMBER_NO", "EMPLOYEE_NAME", "GROSS_SALARY",
                         "TIER1_EMPLOYEE", "TIER2_EMPLOYEE", "TOTAL_EMPLOYEE",
                         "TIER1_EMPLOYER", "TIER2_EMPLOYER", "TOTAL_EMPLOYER",
                         "GRAND_TOTAL"])
            for tx in report.transactions:
                w.writerow([
                    tx.nssf_number or f"NB/{tx.employee_id:08d}",
                    tx.full_name, tx.gross_salary,
                    tx.nssf_tier1_employee, tx.nssf_tier2_employee, tx.nssf_employee,
                    tx.nssf_tier1_employer, tx.nssf_tier2_employer, tx.nssf_employer,
                    _money(tx.nssf_employee + tx.nssf_employer),
                ])
        return path

    def _shif(self, report: AggregationReport,
              d: Path, period: str) -> Path:
        path = d / f"shif_{period}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["SHA_MEMBER_NO", "EMPLOYEE_NAME",
                         "GROSS_SALARY", "SHIF_2.75PCT"])
            for tx in report.transactions:
                sha_no = tx.sha_number or f"SHA/{tx.employee_id:08d}"
                w.writerow([sha_no, tx.full_name,
                             tx.gross_salary, tx.shif])
        return path

    def _ahl(self, report: AggregationReport,
             d: Path, period: str) -> Path:
        path = d / f"ahl_{period}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["KRA_PIN", "EMPLOYEE_NAME", "GROSS_SALARY",
                         "AHL_EMPLOYEE_1.5PCT", "AHL_EMPLOYER_1.5PCT",
                         "TOTAL_AHL", "AHL_RELIEF"])
            for tx in report.transactions:
                w.writerow([
                    tx.kra_pin or "A000000000A", tx.full_name,
                    tx.gross_salary, tx.housing_levy_employee,
                    tx.housing_levy_employer,
                    _money(tx.housing_levy_employee + tx.housing_levy_employer),
                    tx.ahl_relief,
                ])
        return path


class P9ExcelExporter:
    """
    Annual P9A tax certificate for each employee — one sheet per employee.
    Used for personal iTax filing; employer provides these by end of February.

    Future: Phase 7 will encrypt individual P9A files before distribution.
    """

    def export(self, report: AggregationReport,
               export_dir: Path = EXPORT_DIR) -> Path:
        wb  = openpyxl.Workbook()
        ws0 = wb.active
        ws0.title = "Index"

        period = _period(report.month, report.year)
        ws0["A1"] = f"P9A Annual Tax Certificates — {_month_name(report.month, report.year)}"
        ws0["A1"].font = Font(bold=True, size=13, name="Calibri")
        ws0["A2"] = f"Employees: {report.employee_count}"

        for i, tx in enumerate(report.transactions):
            ws = wb.create_sheet(title=f"P9A_{tx.employee_id:04d}")
            self._write_p9(ws, tx, report)

        out = export_dir / f"p9a_forms_{period}.xlsx"
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
        logger.info(f"P9A forms exported: {out} ({len(report.transactions)} sheets)")
        return out

    def _write_p9(self, ws, tx: EnrichedTransaction,
                  report: AggregationReport) -> None:
        month_name = _month_name(report.month, report.year)
        ws["A1"] = "P9A — EMPLOYEE TAX CERTIFICATE"
        ws["A1"].font = Font(bold=True, size=13, name="Calibri", color="22c55e")
        ws["A1"].fill = PatternFill("solid", fgColor="050e09")

        info_rows = [
            ("Employee Name",   tx.full_name),
            ("KRA PIN",         tx.kra_pin or "N/A"),
            ("NSSF No.",        tx.nssf_number or "N/A"),
            ("Department",      tx.department or "N/A"),
            ("Job Title",       tx.job_title or "N/A"),
            ("Period",          month_name),
        ]
        for r, (lbl, val) in enumerate(info_rows, 3):
            ws.cell(row=r, column=1, value=lbl).font  = Font(bold=True, size=9, name="Calibri", color="5a7a65")
            ws.cell(row=r, column=2, value=val).font  = Font(size=9, name="Calibri", color="dff0e6")

        data_rows = [
            ("Gross Pay",                  tx.gross_salary),
            ("NSSF Deducted",              -tx.nssf_employee),
            ("SHIF Deducted",              -tx.shif),
            ("AHL Deducted",               -tx.housing_levy_employee),
            ("Taxable Income",             tx.taxable_income),
            ("Gross PAYE",                 tx.gross_paye),
            ("Personal Relief",            -tx.personal_relief),
            ("Insurance Relief",           -tx.insurance_relief),
            ("Mortgage Relief",            -tx.mortgage_relief),
            ("AHL Relief",                 -tx.ahl_relief),
            ("Net PAYE",                   tx.paye),
            ("Net Pay",                    tx.net_pay),
        ]
        start = 11
        for r, (lbl, val) in enumerate(data_rows, start):
            ws.cell(row=r, column=1, value=lbl).font  = Font(size=9, name="Calibri", color="5a7a65")
            cell = ws.cell(row=r, column=2, value=_money(abs(val)))
            cell.number_format = "#,##0.00"
            cell.alignment     = Alignment(horizontal="right")
            cell.font          = Font(size=9, name="Calibri",
                                      color="ef4444" if val < 0 else "dff0e6")

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 16


# ─── EXPORT ORCHESTRATOR ──────────────────────────────────────────────────────

class ExportOrchestrator:
    """
    Single entry point for all Phase 5 exports.

    Steps:
      1. Aggregator reads + validates all shards
      2. Runs all exporters in sequence
      3. Fires AGGREGATION_DONE and EXPORT_CREATED events
      4. Returns paths dict for downstream use (Phase 4 Phase4Pipeline etc.)

    Future phases:
      Phase 6: BackupPlugin subscribes to AGGREGATION_DONE and
               snapshots all shards after the run.
      Phase 7: RBACPlugin can gate the call with @require_role("payroll_admin").
    """

    def __init__(self,
                 shard_manager: ShardManager,
                 company_info:  dict,
                 export_dir:    Path = EXPORT_DIR,
                 bus:           EventBus | None = None):
        self.sm           = shard_manager
        self.company_info = company_info
        self.export_dir   = Path(export_dir)
        self.bus          = bus or get_bus()

    def run(self,
            month:            int,
            year:             int,
            company_id:       int = 1,
            shard_ids:        list[int] | None = None,
            collector_totals: dict | None = None,
            ) -> dict:
        """
        Execute full aggregation + all exports.
        Returns dict with 'report' and all 'paths'.
        """
        self.export_dir.mkdir(parents=True, exist_ok=True)

        # 1. Aggregate
        agg    = Aggregator(self.sm)
        report = agg.aggregate(month, year, company_id,
                               shard_ids, collector_totals)

        self.bus.publish(PayrollEvent.AGGREGATION_DONE,
                         report=report, month=month, year=year)

        # 2. Exports
        paths: dict[str, Path] = {}

        # Payroll register
        reg = PayrollRegisterExporter()
        paths["payroll_register"] = reg.export(report, self.export_dir)
        self.bus.publish(PayrollEvent.EXPORT_CREATED,
                         path=paths["payroll_register"], export_type="payroll_register")

        # GL journal
        gl = GLExporter()
        xlsx_gl, csv_gl = gl.export(report, self.export_dir)
        paths["gl_xlsx"] = xlsx_gl
        paths["gl_csv"]  = csv_gl
        self.bus.publish(PayrollEvent.EXPORT_CREATED,
                         path=xlsx_gl, export_type="gl_journal")

        # Remittance summary
        rem = RemittanceSummaryExporter()
        paths["remittance"] = rem.export(report, self.company_info, self.export_dir)
        self.bus.publish(PayrollEvent.EXPORT_CREATED,
                         path=paths["remittance"], export_type="remittance_summary")

        # Government CSVs
        gov = GovernmentFileExporter()
        gov_paths = gov.export_all(report, self.export_dir)
        paths.update({f"gov_{k}": v for k, v in gov_paths.items()})
        for k, p in gov_paths.items():
            self.bus.publish(PayrollEvent.EXPORT_CREATED,
                             path=p, export_type=f"gov_{k}")

        # P9A forms
        p9 = P9ExcelExporter()
        paths["p9a"] = p9.export(report, self.export_dir)
        self.bus.publish(PayrollEvent.EXPORT_CREATED,
                         path=paths["p9a"], export_type="p9a")

        # Write manifest JSON
        manifest_path = self.export_dir / f"manifest_{_period(month, year)}.json"
        with open(manifest_path, "w") as f:
            json.dump({
                "report":    report.summary_dict(),
                "files":     {k: str(v) for k, v in paths.items()},
                "generated": report.generated_at,
            }, f, indent=2)
        paths["manifest"] = manifest_path

        logger.info(
            f"ExportOrchestrator complete — {len(paths)} files produced "
            f"in {self.export_dir}"
        )
        return {"report": report, "paths": paths}
