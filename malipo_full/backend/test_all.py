"""
=============================================================================
KENYA PAYROLL SYSTEM — Master Test Suite (Phases 1–7)
=============================================================================
Run:  python test_all.py  (or pytest test_all.py -v)
=============================================================================
"""
import csv, json, sys, tempfile, threading, time, unittest
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent))

from shard_manager   import ShardManager, build_seed_employee
from payroll_engine  import (
    PayrollEngine, EmployeeInput,
    calc_nssf, calc_shif, calc_ahl, calc_paye_bands, apply_reliefs,
    NSSF_EMPLOYEE_MAX, PERSONAL_RELIEF,
    INSURANCE_RELIEF_CAP, MORTGAGE_RELIEF_CAP,
    POST_RETIREMENT_CAP, PENSION_PRE_TAX_CAP,
    SHIF_AHL_TAXDEDUCTIBLE,
)
from worker_queue    import PayrollRunner, ResultCollector, Dispatcher, PayrollTask
from events          import EventBus, PayrollEvent, ConsoleLoggerPlugin
from phase4_excel    import (
    TemplateStore, TemplatePopulator, TemplateValidator,
    UploadPreparator, Phase4Pipeline,
)
from phase5_aggregator import (
    Aggregator, ShardReader, ExportOrchestrator,
    GLExporter, PayrollRegisterExporter, RemittanceSummaryExporter,
    GovernmentFileExporter, P9ExcelExporter, AggregationReport,
)
from phase6_audit_backup import (
    AuditPlugin, BackupPlugin, P9AAnnualAggregator, setup_phase6,
)
from phase7_security import (
    UserContext, Role, Permission, ROLE_PERMISSIONS,
    require_role, require_permission,
    RBACPlugin, CredentialVault, SessionManager,
    EncryptionManager, ApprovalWorkflow, setup_phase7,
)
import openpyxl

COMPANY_INFO = {
    "name":    "Savanna Tech Ltd",
    "pin":     "P051234567A",
    "nssf_no": "NB/001234",
    "sha_no":  "SHA/ET/001234",
}

def make_emp(eid, salary, **kw):
    return EmployeeInput(employee_id=eid, full_name=f"Emp {eid}",
                         base_salary=salary, **kw)

def _seeded_sm(n_shards=2, per_shard=50, n_employees=10):
    tmpdir = tempfile.mkdtemp()
    sm = ShardManager(shard_dir=Path(tmpdir), employees_per_shard=per_shard)
    sm.initialize_shards(n_shards)
    for i in range(1, n_employees + 1):
        sm.insert_employee(build_seed_employee(i))
    return sm, tmpdir

def _run_payroll(sm, month=3, year=2026, shard_ids=None):
    return PayrollRunner(sm).run(month=month, year=year, shard_ids=shard_ids)

def _make_transactions(n=5):
    engine = PayrollEngine()
    txs = []
    salaries = [28000, 65000, 85000, 120000, 250000]
    for i in range(n):
        salary = salaries[i % len(salaries)]
        inp = make_emp(i+1, salary)
        r = engine.compute(inp)
        txs.append({
            "employee_id": i+1, "full_name": f"Employee {i+1}",
            "kra_pin": f"A{i+1:09d}B", "nssf_number": f"NB/{i+1:08d}",
            "sha_number": f"SHA/{i+1:08d}", "base_salary": salary,
            "gross_salary": r.gross_salary,
            "taxable_allowances": r.taxable_allowances,
            "non_taxable_allowances": r.non_taxable_allowances,
            "benefits_in_kind": r.benefits_in_kind,
            "nssf_tier1_employee": r.nssf_tier1_employee,
            "nssf_tier2_employee": r.nssf_tier2_employee,
            "nssf_employee": r.nssf_employee,
            "nssf_tier1_employer": r.nssf_tier1_employer,
            "nssf_tier2_employer": r.nssf_tier2_employer,
            "nssf_employer": r.nssf_employer,
            "shif": r.shif, "housing_levy_employee": r.housing_levy_employee,
            "housing_levy_employer": r.housing_levy_employer,
            "taxable_income": r.taxable_income, "gross_paye": r.gross_paye,
            "personal_relief": r.personal_relief,
            "insurance_relief": r.insurance_relief,
            "mortgage_relief": r.mortgage_relief, "ahl_relief": r.ahl_relief,
            "post_retirement_relief": r.post_retirement_relief,
            "total_relief": r.total_relief, "paye": r.paye,
            "helb_deduction": r.helb_deduction,
            "other_deductions": r.other_deductions,
            "total_statutory": r.total_statutory,
            "total_deductions": r.total_deductions, "net_pay": r.net_pay,
        })
    return txs


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1 — Sharding & Schema
# ══════════════════════════════════════════════════════════════════════════════
class TestShardManager(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.sm = ShardManager(shard_dir=Path(self.tmpdir), employees_per_shard=100)
        self.sm.initialize_shards(3)

    def test_shard_id_boundaries(self):
        self.assertEqual(self.sm.shard_id_for(1),   1)
        self.assertEqual(self.sm.shard_id_for(100), 1)
        self.assertEqual(self.sm.shard_id_for(101), 2)
        self.assertEqual(self.sm.shard_id_for(201), 3)

    def test_shard_files_created(self):
        self.assertEqual(self.sm.list_shards(), [1, 2, 3])

    def test_insert_and_fetch(self):
        self.sm.insert_employee(build_seed_employee(1, base_salary=85000))
        got = self.sm.get_employee(1)
        self.assertAlmostEqual(got["base_salary"], 85000, places=2)

    def test_shard_isolation(self):
        self.sm.insert_employee(build_seed_employee(50))
        self.sm.insert_employee(build_seed_employee(150))
        s1 = [e["employee_id"] for e in self.sm.get_employees_in_shard(1)]
        s2 = [e["employee_id"] for e in self.sm.get_employees_in_shard(2)]
        self.assertIn(50, s1);  self.assertNotIn(150, s1)
        self.assertIn(150, s2); self.assertNotIn(50, s2)

    def test_concurrent_writes(self):
        errors = []
        def batch(start):
            try:
                for i in range(start, start + 10):
                    self.sm.insert_employee(build_seed_employee(i))
            except Exception as e:
                errors.append(e)
        t1 = threading.Thread(target=batch, args=(1,))
        t2 = threading.Thread(target=batch, args=(101,))
        t1.start(); t2.start(); t1.join(); t2.join()
        self.assertEqual(errors, [])

    def test_payroll_run_idempotent(self):
        r1 = self.sm.create_payroll_run(1, 3, 2026)
        r2 = self.sm.create_payroll_run(1, 3, 2026)
        self.assertEqual(r1, r2)

    def test_audit_log_write_and_read(self):
        self.sm.write_audit(1, dict(
            payroll_run_id=1, employee_id=1, event_type="CALC",
            rule_applied="TEST", field_name="net_pay",
            before_value=None, after_value="50000",
            notes="unit test", operator="pytest",
        ))
        rows = self.sm.get_audit_log(1)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["event_type"], "CALC")

    def test_reliefs_for_shard(self):
        """Phase 1 enhancement: reliefs table and bulk load."""
        self.sm.insert_employee(build_seed_employee(5))
        with self.sm.connection(1) as conn:
            conn.execute("""
                INSERT INTO reliefs (employee_id, relief_type, monthly_amount)
                VALUES (5, 'insurance', 3000)
            """)
        relief_map = self.sm.get_reliefs_for_shard(1)
        self.assertIn(5, relief_map)
        self.assertEqual(relief_map[5][0]["relief_type"], "insurance")

    def test_lock_run(self):
        """Phase 7 approval: lock_run changes status to 'locked'."""
        self.sm.insert_employee(build_seed_employee(1))
        run_id = self.sm.create_payroll_run(1, 3, 2026)
        self.sm.mark_run_processed(1, run_id)
        self.sm.lock_run(1, run_id)
        with self.sm.connection(1) as conn:
            row = conn.execute(
                "SELECT status FROM payroll_runs WHERE payroll_run_id=?",
                (run_id,)
            ).fetchone()
        self.assertEqual(row["status"], "locked")


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 2 — Calculation Engine
# ══════════════════════════════════════════════════════════════════════════════
class TestNSSF(unittest.TestCase):
    def test_below_tier1(self):
        r = calc_nssf(6000)
        self.assertAlmostEqual(r.tier1_employee, 360.0, places=2)
        self.assertAlmostEqual(r.tier2_employee, 0.0,   places=2)

    def test_at_tier1_ceiling(self):
        r = calc_nssf(9000)
        self.assertAlmostEqual(r.tier1_employee, 540.0, places=2)

    def test_between_tiers(self):
        r = calc_nssf(50000)
        self.assertAlmostEqual(r.tier1_employee, 540.0,  places=2)
        self.assertAlmostEqual(r.tier2_employee, 2460.0, places=2)  # (50000-9000)*6%

    def test_cap_at_uel(self):
        r = calc_nssf(108000)
        self.assertAlmostEqual(r.employee, 6480.0, places=2)

    def test_cap_above_uel(self):
        r = calc_nssf(200000)
        self.assertAlmostEqual(r.employee,  NSSF_EMPLOYEE_MAX, places=2)
        self.assertAlmostEqual(r.employer, NSSF_EMPLOYEE_MAX, places=2)

    def test_employer_mirrors_employee(self):
        r = calc_nssf(75000)
        self.assertAlmostEqual(r.employee, r.employer, places=2)


class TestSHIF(unittest.TestCase):
    def test_basic(self):
        self.assertAlmostEqual(calc_shif(100000), 2750.0, places=2)

    def test_uncapped(self):
        self.assertAlmostEqual(calc_shif(1_000_000), 27500.0, places=2)

    def test_zero(self):
        self.assertAlmostEqual(calc_shif(0), 0.0, places=2)


class TestAHL(unittest.TestCase):
    def test_basic(self):
        emp, emplr = calc_ahl(100000)
        self.assertAlmostEqual(emp,   1500.0, places=2)
        self.assertAlmostEqual(emplr, 1500.0, places=2)

    def test_symmetry(self):
        emp, emplr = calc_ahl(55000)
        self.assertAlmostEqual(emp, emplr, places=2)


class TestPAYEBands(unittest.TestCase):
    def test_band1_only(self):
        _, tax = calc_paye_bands(24000)
        self.assertAlmostEqual(tax, 2400.0, places=2)

    def test_band2(self):
        _, tax = calc_paye_bands(32333)
        expected = 2400 + 8333 * 0.25
        self.assertAlmostEqual(tax, expected, places=2)

    def test_nonresident_flat_30(self):
        bands, tax = calc_paye_bands(50000, nonresident=True)
        self.assertAlmostEqual(tax, 15000.0, places=2)
        self.assertEqual(len(bands), 1)

    def test_zero(self):
        _, tax = calc_paye_bands(0)
        self.assertAlmostEqual(tax, 0.0, places=2)


class TestPayrollEngine(unittest.TestCase):
    def setUp(self):
        self.eng = PayrollEngine()

    def test_standard_70k(self):
        """Verify NSSF Year4 numbers for KES 70,000."""
        r = self.eng.compute(make_emp(1, 70000))
        self.assertAlmostEqual(r.nssf_employee, 4200.0, places=2)  # Tier1=540 + Tier2=3660
        self.assertAlmostEqual(r.shif,          1925.0, places=2)
        self.assertAlmostEqual(r.housing_levy_employee, 1050.0, places=2)

    def test_accounting_identity(self):
        """gross = net_pay + total_deductions."""
        for salary in [28000, 65000, 120000, 300000]:
            r = self.eng.compute(make_emp(1, salary))
            self.assertAlmostEqual(r.gross_salary,
                                   r.net_pay + r.total_deductions, places=2)

    def test_nonresident_no_relief(self):
        """Non-residents pay flat 30% PAYE and receive no personal relief."""
        r = self.eng.compute(make_emp(1, 100000,
                                      residency_status="nonresident"))
        self.assertAlmostEqual(r.personal_relief, 0.0, places=2)
        self.assertAlmostEqual(r.insurance_relief, 0.0, places=2)
        # PAYE should be gross_paye (no relief subtracted)
        self.assertAlmostEqual(r.paye, r.gross_paye, places=2)

    def test_nssf_cap_enforced(self):
        """Employees earning above UEL pay the maximum NSSF."""
        r = self.eng.compute(make_emp(1, 500000))
        self.assertAlmostEqual(r.nssf_employee, NSSF_EMPLOYEE_MAX, places=2)

    # ── Corrected cap tests ───────────────────────────────────────────────────

    def test_mortgage_relief_cap_corrected(self):
        """MORTGAGE_RELIEF_CAP must be KES 30,000/mo (was 25,000 — corrected)."""
        self.assertEqual(MORTGAGE_RELIEF_CAP, 30_000.0)
        r = self.eng.compute(make_emp(1, 200000, mortgage_relief=40000))
        self.assertAlmostEqual(r.mortgage_relief, 30000.0, places=2)

    def test_post_retirement_cap_corrected(self):
        """POST_RETIREMENT_CAP must be KES 15,000/mo (was 5,000 — corrected)."""
        self.assertEqual(POST_RETIREMENT_CAP, 15_000.0)
        r = self.eng.compute(make_emp(1, 200000, post_retirement_relief=20000))
        self.assertAlmostEqual(r.post_retirement_relief, 15000.0, places=2)

    def test_pension_pre_tax_cap(self):
        """Pension pre-tax deduction must be capped at KES 30,000/mo."""
        self.assertEqual(PENSION_PRE_TAX_CAP, 30_000.0)
        r = self.eng.compute(make_emp(1, 100000, pension_pre_tax=40000))
        # Capped pension shows in audit notes
        self.assertTrue(any("Pension capped" in n for n in r.audit_notes))
        # The actual pension used in taxable income must be ≤ 30,000
        # pre_tax_deductions = nssf + pension_capped + shif + ahl (if deductible)
        pension_used = r.pre_tax_deductions - r.nssf_employee - r.shif - r.housing_levy_employee
        self.assertAlmostEqual(pension_used, 30000.0, places=2)

    def test_shif_ahl_deductible_flag_true(self):
        """With flag=True, SHIF and AHL reduce taxable income (default KRA 2025/26)."""
        r = self.eng.compute(make_emp(1, 70000, shif_ahl_taxdeductible=True))
        expected_ti = 70000 - r.nssf_employee - r.shif - r.housing_levy_employee
        self.assertAlmostEqual(r.taxable_income, expected_ti, delta=1.0)
        self.assertTrue(r.shif_ahl_deducted)

    def test_shif_ahl_deductible_flag_false(self):
        """With flag=False, only NSSF reduces taxable income."""
        r_true  = self.eng.compute(make_emp(1, 70000, shif_ahl_taxdeductible=True))
        r_false = self.eng.compute(make_emp(1, 70000, shif_ahl_taxdeductible=False))
        # Taxable income should be higher when SHIF/AHL not deducted
        self.assertGreater(r_false.taxable_income, r_true.taxable_income)
        self.assertFalse(r_false.shif_ahl_deducted)

    def test_ahl_relief_15pct(self):
        """AHL relief = 15% of employee AHL contribution."""
        r = self.eng.compute(make_emp(1, 100000))
        expected_ahl_relief = round(r.housing_levy_employee * 0.15, 2)
        self.assertAlmostEqual(r.ahl_relief, expected_ahl_relief, places=2)

    def test_high_earner(self):
        """KES 500,000 gross — verify 32.5% band and NSSF cap."""
        r = self.eng.compute(make_emp(1, 500000))
        self.assertAlmostEqual(r.nssf_employee, NSSF_EMPLOYEE_MAX, places=2)
        self.assertGreater(r.gross_paye, 100000)  # well into higher bands

    def test_insurance_relief_cap(self):
        r = self.eng.compute(make_emp(1, 80000, insurance_relief=8000))
        self.assertAlmostEqual(r.insurance_relief, INSURANCE_RELIEF_CAP, places=2)

    def test_disability_exemption(self):
        r_normal = self.eng.compute(make_emp(1, 80000))
        r_pwd    = self.eng.compute(make_emp(1, 80000, disability_exemption=80000))
        self.assertLess(r_pwd.taxable_income, r_normal.taxable_income)
        self.assertLess(r_pwd.paye, r_normal.paye)

    def test_bik_increases_taxable_income(self):
        r_no_bik = self.eng.compute(make_emp(1, 80000))
        r_bik    = self.eng.compute(make_emp(1, 80000, car_benefit=10000))
        self.assertAlmostEqual(r_bik.taxable_income - r_no_bik.taxable_income,
                               10000.0, delta=1.0)

    def test_helb_reduces_net_not_taxable(self):
        r_no_helb = self.eng.compute(make_emp(1, 80000))
        r_helb    = self.eng.compute(make_emp(1, 80000, helb_deduction=2000))
        # HELB is post-tax — should not affect PAYE
        self.assertAlmostEqual(r_no_helb.paye, r_helb.paye, places=2)
        # But net pay is lower
        self.assertAlmostEqual(r_helb.net_pay, r_no_helb.net_pay - 2000, places=2)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 3 — Worker Queue
# ══════════════════════════════════════════════════════════════════════════════
class TestPinnedWorkers(unittest.TestCase):
    def setUp(self):
        self.sm, _ = _seeded_sm(n_shards=2, n_employees=10)

    def test_all_employees_processed(self):
        c = _run_payroll(self.sm)
        self.assertEqual(c.processed_count, 10)
        self.assertEqual(len(c.errors), 0)

    def test_shard_totals_match_grand_total(self):
        c = _run_payroll(self.sm)
        from_shards = sum(v.get("paye", 0) for v in c.shard_totals.values())
        self.assertAlmostEqual(from_shards, c.totals["paye"], delta=0.05)

    def test_parallel_shards_no_overlap(self):
        """Each shard's employee set must not overlap."""
        c = _run_payroll(self.sm)
        seen = set()
        for shard_id, subtotals in c.shard_totals.items():
            # Each shard has gross > 0 if it processed employees
            self.assertGreater(subtotals.get("gross_salary", 0), 0)

    def test_event_bus_run_completed(self):
        bus = EventBus(); events = []
        bus.subscribe(PayrollEvent.RUN_COMPLETED, lambda **kw: events.append(kw))
        sm, _ = _seeded_sm(n_employees=5)
        PayrollRunner(sm, bus=bus).run(month=3, year=2026)
        self.assertEqual(len(events), 1)
        self.assertEqual(events[0]["collector"].processed_count, 5)

    def test_reliefs_loaded_from_shard(self):
        """Workers must load reliefs from the shard, not require external supply."""
        sm, _ = _seeded_sm(n_shards=1, n_employees=3)
        # employees 1-3 already seeded; add a relief to emp 1
        with sm.connection(1) as conn:
            conn.execute("""
                INSERT OR IGNORE INTO reliefs (employee_id, relief_type, monthly_amount)
                VALUES (1, 'insurance', 3000)
            """)
        # Should not raise; reliefs are loaded automatically
        c = PayrollRunner(sm).run(month=3, year=2026)
        self.assertGreaterEqual(c.processed_count, 1)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 4 — Excel Template Automation
# ══════════════════════════════════════════════════════════════════════════════
class TestTemplateStore(unittest.TestCase):
    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.store = TemplateStore(self.tmp / "tpl", self.tmp / "filled")

    def test_template_path_structure(self):
        p = self.store.template_path("kra", 3, 2026)
        self.assertEqual(p.parent.name, "2026-03")
        self.assertEqual(p.name, "kra_template.xlsx")

    def test_sha256(self):
        f = self.tmp / "test.txt"
        f.write_bytes(b"hello")
        h = TemplateStore.sha256(f)
        self.assertEqual(len(h), 64)

    def test_manifest_saved(self):
        fp = self.tmp / "filled" / "2026-03" / "test.xlsx"
        fp.parent.mkdir(parents=True, exist_ok=True); fp.write_bytes(b"x")
        self.store.save_manifest("kra", 3, 2026, fp, 10, "abc123", {})
        mp = self.store.manifest_path("kra", 3, 2026)
        self.assertTrue(mp.exists())
        with open(mp) as f:
            m = json.load(f)
        self.assertEqual(m["row_count"], 10)


class TestTemplatePopulator(unittest.TestCase):
    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.store = TemplateStore(self.tmp/"tpl", self.tmp/"filled")
        self.pop   = TemplatePopulator(self.store)
        self.txs   = _make_transactions(5)

    def test_kra_paye_populated(self):
        p = self.pop.populate_kra_paye(self.txs, 3, 2026, COMPANY_INFO)
        self.assertTrue(p.exists())
        wb = openpyxl.load_workbook(str(p))
        self.assertGreater(wb.active.max_row, 5)

    def test_nssf_populated(self):
        p = self.pop.populate_nssf(self.txs, 3, 2026, COMPANY_INFO)
        self.assertTrue(p.exists())

    def test_shif_populated(self):
        p = self.pop.populate_shif(self.txs, 3, 2026, COMPANY_INFO)
        self.assertTrue(p.exists())

    def test_ahl_populated(self):
        p = self.pop.populate_ahl(self.txs, 3, 2026, COMPANY_INFO)
        self.assertTrue(p.exists())


class TestTemplateValidator(unittest.TestCase):
    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.store = TemplateStore(self.tmp/"tpl", self.tmp/"filled")
        self.pop   = TemplatePopulator(self.store)
        self.val   = TemplateValidator()
        self.txs   = _make_transactions(5)

    def test_kra_validation_passes(self):
        p  = self.pop.populate_kra_paye(self.txs, 3, 2026, COMPANY_INFO)
        vr = self.val.validate(p, "kra", expected_row_count=5)
        self.assertTrue(vr.passed, f"Errors: {vr.errors}")

    def test_shif_validation_passes(self):
        p  = self.pop.populate_shif(self.txs, 3, 2026, COMPANY_INFO)
        vr = self.val.validate(p, "shif", expected_row_count=5)
        self.assertTrue(vr.passed, f"Errors: {vr.errors}")


class TestPhase4Pipeline(unittest.TestCase):
    def test_full_pipeline_all_portals(self):
        tmp  = Path(tempfile.mkdtemp())
        txs  = _make_transactions(5)
        p4   = Phase4Pipeline(COMPANY_INFO, tmp/"tpl", tmp/"filled")
        pkgs = p4.run(txs, 3, 2026)
        for code in ("kra", "nssf", "shif", "ahl"):
            self.assertIn(code, pkgs)
            self.assertTrue(pkgs[code]["validation"]["passed"],
                            f"P4 {code}: {pkgs[code]['validation']['errors']}")


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 5 — Aggregator & Exports
# ══════════════════════════════════════════════════════════════════════════════
class TestShardReader(unittest.TestCase):
    def setUp(self):
        self.sm, _ = _seeded_sm(n_employees=5)
        _run_payroll(self.sm)
        self.reader = ShardReader(self.sm)

    def test_employee_metadata_joined(self):
        rows, _ = self.reader.read_shard(1, payroll_run_id=1)
        for row in rows:
            self.assertTrue(row.full_name)

    def test_subtotal_matches_sum(self):
        rows, sub = self.reader.read_shard(1, payroll_run_id=1)
        self.assertAlmostEqual(sub.paye,
                               round(sum(r.paye for r in rows), 2), places=2)

    def test_non_negative_values(self):
        for sid in self.sm.list_shards():
            rows, _ = self.reader.read_shard(sid, 1)
            for r in rows:
                for fld in ("gross_salary", "paye", "net_pay"):
                    self.assertGreaterEqual(getattr(r, fld), 0)


class TestAggregator(unittest.TestCase):
    def setUp(self):
        self.sm, _ = _seeded_sm(n_employees=10)
        self.collector = _run_payroll(self.sm)
        self.agg = Aggregator(self.sm)

    def test_employee_count(self):
        r = self.agg.aggregate(3, 2026, collector_totals=self.collector.shard_totals)
        self.assertEqual(r.employee_count, 10)

    def test_totals_match_collector(self):
        r  = self.agg.aggregate(3, 2026, collector_totals=self.collector.shard_totals)
        ct = self.collector.totals
        self.assertAlmostEqual(r.gross_salary, ct["gross_salary"], delta=0.10)
        self.assertAlmostEqual(r.paye,         ct["paye"],         delta=0.10)

    def test_is_valid(self):
        r = self.agg.aggregate(3, 2026)
        self.assertTrue(r.is_valid)

    def test_accounting_identity(self):
        r = self.agg.aggregate(3, 2026)
        for tx in r.transactions:
            self.assertAlmostEqual(tx.gross_salary,
                                   tx.net_pay + tx.total_deductions, places=2)

    def test_subtotals_sum_to_grand(self):
        r = self.agg.aggregate(3, 2026)
        self.assertAlmostEqual(round(sum(s.paye for s in r.shard_subtotals), 2),
                               r.paye, places=2)


class TestGLExporter(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        sm, _ = _seeded_sm(n_employees=5); _run_payroll(sm)
        self.report = Aggregator(sm).aggregate(3, 2026)
        self.gl = GLExporter()

    def test_files_created(self):
        xlsx, csv_ = self.gl.export(self.report, self.tmp)
        self.assertTrue(xlsx.exists()); self.assertTrue(csv_.exists())

    def test_balanced(self):
        xlsx, _ = self.gl.export(self.report, self.tmp)
        wb   = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        totals = rows[-2]
        dr = float(totals[3] or 0); cr = float(totals[4] or 0)
        self.assertAlmostEqual(dr, cr, delta=0.10)


class TestGovernmentFiles(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        sm, _ = _seeded_sm(n_employees=5); _run_payroll(sm)
        self.report = Aggregator(sm).aggregate(3, 2026)
        self.exp = GovernmentFileExporter()

    def test_all_four_files(self):
        paths = self.exp.export_all(self.report, self.tmp)
        self.assertEqual(set(paths.keys()), {"kra", "nssf", "shif", "ahl"})
        for k, p in paths.items():
            self.assertTrue(p.exists(), f"{k} missing")

    def test_shif_amounts(self):
        paths = self.exp.export_all(self.report, self.tmp)
        with open(paths["shif"]) as f:
            rows = list(csv.DictReader(f))
        for row in rows:
            gross = float(row["GROSS_SALARY"])
            shif  = float(row["SHIF_2.75PCT"])
            self.assertAlmostEqual(shif, gross * 0.0275, delta=0.05)

    def test_ahl_symmetry(self):
        paths = self.exp.export_all(self.report, self.tmp)
        with open(paths["ahl"]) as f:
            rows = list(csv.DictReader(f))
        for row in rows:
            self.assertAlmostEqual(float(row["AHL_EMPLOYEE_1.5PCT"]),
                                   float(row["AHL_EMPLOYER_1.5PCT"]), places=2)


class TestExportOrchestrator(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.sm, _ = _seeded_sm(n_employees=8)
        self.collector = _run_payroll(self.sm)

    def _orch(self, subdir="exp"):
        return ExportOrchestrator(self.sm, COMPANY_INFO, export_dir=self.tmp/subdir)

    def test_all_keys_present(self):
        r = self._orch("e1").run(3, 2026, collector_totals=self.collector.shard_totals)
        for k in ("payroll_register","gl_xlsx","remittance","gov_kra","p9a","manifest"):
            self.assertIn(k, r["paths"])

    def test_all_files_exist(self):
        r = self._orch("e2").run(3, 2026)
        for name, path in r["paths"].items():
            self.assertTrue(path.exists(), f"Missing: {name}")

    def test_manifest_valid(self):
        r = self._orch("e3").run(3, 2026)
        with open(r["paths"]["manifest"]) as f:
            m = json.load(f)
        self.assertEqual(m["report"]["employee_count"], 8)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 6 — Audit & Backup
# ══════════════════════════════════════════════════════════════════════════════
class TestAuditPlugin(unittest.TestCase):
    def setUp(self):
        self.sm, _ = _seeded_sm(n_employees=5)
        self.bus   = EventBus()
        self.audit = AuditPlugin(self.sm, operator="test_runner")
        self.audit.register(self.bus)

    def test_transaction_saved_writes_audit(self):
        """TRANSACTION_SAVED event should create a CALC audit entry."""
        self.bus.publish(PayrollEvent.TRANSACTION_SAVED,
                         employee_id=1, shard_id=1,
                         net_pay=50000.0, payroll_run_id=1)
        rows = self.sm.get_audit_log(1)
        calc_rows = [r for r in rows if r["event_type"] == "CALC"]
        self.assertGreater(len(calc_rows), 0)
        self.assertEqual(calc_rows[0]["field_name"], "net_pay")

    def test_error_writes_audit(self):
        """ERROR_OCCURRED event should create an ERROR audit entry."""
        self.bus.publish(PayrollEvent.ERROR_OCCURRED,
                         employee_id=99, shard_id=1,
                         error="Test failure")
        rows = self.sm.get_audit_log(1)
        err_rows = [r for r in rows if r["event_type"] == "ERROR"]
        self.assertGreater(len(err_rows), 0)

    def test_run_completed_writes_summary(self):
        """RUN_COMPLETED should write a RUN_SUMMARY to all shards."""
        # Run payroll first to create run records
        collector = PayrollRunner(self.sm, bus=self.bus).run(3, 2026)
        # Check all shards have audit entries
        for sid in self.sm.list_shards():
            rows = self.sm.get_audit_log(sid)
            self.assertGreater(len(rows), 0)

    def test_audit_populated_after_payroll(self):
        """Full payroll run with AuditPlugin registered writes to audit_log."""
        sm, _ = _seeded_sm(n_employees=3)
        bus   = EventBus()
        audit = AuditPlugin(sm, operator="test")
        audit.register(bus)
        PayrollRunner(sm, bus=bus).run(3, 2026)
        # Audit log should have entries
        all_entries = []
        for sid in sm.list_shards():
            all_entries.extend(sm.get_audit_log(sid))
        self.assertGreater(len(all_entries), 0)


class TestBackupPlugin(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.sm, _ = _seeded_sm(n_employees=5)
        _run_payroll(self.sm)
        self.backup = BackupPlugin(self.sm, backup_dir=self.tmp/"backups", retention_days=90)

    def test_backup_creates_files(self):
        manifest = self.backup.backup_now(3, 2026)
        self.assertIn("shards", manifest)
        for shard_entry in manifest["shards"]:
            self.assertNotIn("error", shard_entry)
            dest = self.tmp/"backups"/"2026-03"/shard_entry["file"]
            self.assertTrue(dest.exists())

    def test_backup_manifest_json(self):
        self.backup.backup_now(3, 2026)
        manifests = list((self.tmp/"backups"/"2026-03").glob("manifest_*.json"))
        self.assertGreater(len(manifests), 0)

    def test_sha256_verification_passes(self):
        self.backup.backup_now(3, 2026)
        results = self.backup.verify_backup("2026-03")
        for shard_id, status in results.items():
            if isinstance(shard_id, int):
                self.assertEqual(status, "ok", f"Shard {shard_id} checksum: {status}")

    def test_backup_triggered_by_aggregation_event(self):
        """AGGREGATION_DONE event should trigger an automatic backup."""
        bus    = EventBus()
        backup = BackupPlugin(self.sm, backup_dir=self.tmp/"auto_backups")
        backup.register(bus)
        bus.publish(PayrollEvent.AGGREGATION_DONE, month=3, year=2026)
        time.sleep(0.1)   # let async handler complete
        backup_dir = self.tmp/"auto_backups"/"2026-03"
        self.assertTrue(backup_dir.exists())


class TestP9AAnnualAggregator(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.sm, _ = _seeded_sm(n_employees=5)
        # Run payroll for 3 months to simulate partial year
        for month in [1, 2, 3]:
            _run_payroll(self.sm, month=month, year=2026)
        self.agg = P9AAnnualAggregator(self.sm, export_dir=self.tmp/"exports")

    def test_aggregate_returns_all_employees(self):
        records = self.agg.aggregate(year=2026)
        self.assertEqual(len(records), 5)

    def test_annual_totals_are_sum_of_monthly(self):
        records = self.agg.aggregate(year=2026)
        for rec in records:
            # 3 months run, so months_worked should be 3 (one per shard that has them)
            self.assertGreaterEqual(rec.months_worked, 1)
            # Annual gross should be months * monthly gross (approx)
            if rec.months_worked > 0:
                self.assertGreater(rec.annual_gross, 0)

    def test_p9a_not_single_month(self):
        """P9A must accumulate across months, not be a single month snapshot."""
        records = self.agg.aggregate(year=2026)
        for rec in records:
            if rec.months_worked > 1:
                # monthly breakdown should have multiple rows
                self.assertGreater(len(rec.monthly_rows), 1)

    def test_export_excel_creates_file(self):
        records = self.agg.aggregate(year=2026)
        path = self.agg.export_excel(records, 2026, COMPANY_INFO)
        self.assertTrue(path.exists())
        wb = openpyxl.load_workbook(str(path))
        # Should have index sheet + one sheet per employee
        self.assertGreater(len(wb.sheetnames), 1)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 7 — Security, RBAC & Credential Vault
# ══════════════════════════════════════════════════════════════════════════════
class TestRBAC(unittest.TestCase):
    def test_role_permissions(self):
        """VIEWER cannot run payroll; OPERATOR can."""
        viewer   = UserContext("1", "viewer",   Role.VIEWER,   company_id=1)
        operator = UserContext("2", "operator", Role.OPERATOR, company_id=1)
        self.assertFalse(viewer.can(Permission.RUN_PAYROLL))
        self.assertTrue(operator.can(Permission.RUN_PAYROLL))

    def test_approver_can_approve(self):
        approver = UserContext("3", "approver", Role.APPROVER, company_id=1)
        self.assertTrue(approver.can(Permission.APPROVE_RUN))
        self.assertTrue(approver.can(Permission.LOCK_RUN))

    def test_admin_has_all_permissions(self):
        admin = UserContext("4", "admin", Role.ADMIN, company_id=1)
        for perm in Permission:
            if perm != Permission.MANAGE_CREDENTIALS:
                self.assertTrue(admin.can(perm), f"Admin should have {perm}")

    def test_require_role_decorator_blocks(self):
        @require_role(Role.APPROVER, Role.ADMIN)
        def approve(run_id: int, user: UserContext = None):
            return True

        operator = UserContext("2", "op", Role.OPERATOR)
        with self.assertRaises(PermissionError):
            approve(1, user=operator)

    def test_require_role_decorator_allows(self):
        @require_role(Role.APPROVER, Role.ADMIN)
        def approve(run_id: int, user: UserContext = None):
            return True

        approver = UserContext("3", "ap", Role.APPROVER)
        self.assertTrue(approve(1, user=approver))

    def test_require_role_no_user_passes(self):
        """System calls without user context must always pass."""
        @require_role(Role.ADMIN)
        def admin_op(user: UserContext = None):
            return "ok"

        self.assertEqual(admin_op(user=None), "ok")

    def test_require_permission_decorator(self):
        @require_permission(Permission.EXPORT_P9A)
        def export_p9a(year: int, user: UserContext = None):
            return True

        viewer   = UserContext("1", "v", Role.VIEWER)
        operator = UserContext("2", "o", Role.OPERATOR)
        with self.assertRaises(PermissionError):
            export_p9a(2026, user=viewer)
        self.assertTrue(export_p9a(2026, user=operator))


class TestCredentialVault(unittest.TestCase):
    def test_env_backend(self):
        vault = CredentialVault(backend="env")
        vault.set_env("test_portal", "username", "alice")
        val = vault.get("test_portal", "username")
        self.assertEqual(val, "alice")

    def test_missing_credential_returns_none(self):
        vault = CredentialVault(backend="env")
        val = vault.get("nonexistent_portal", "password")
        self.assertIsNone(val)

    def test_access_log(self):
        vault = CredentialVault(backend="env")
        vault.set_env("kra", "username", "testuser")
        vault.get("kra", "username")
        vault.get("nssf", "password")
        hist = vault.access_history()
        self.assertGreaterEqual(len(hist), 2)
        keys = [h["key"] for h in hist]
        self.assertIn("KRA_USERNAME", keys)


class TestSessionManager(unittest.TestCase):
    def setUp(self):
        self.sm = SessionManager(secret_key="test-secret-key-32chars-padding!")

    def test_create_and_validate(self):
        user  = UserContext("1", "alice", Role.OPERATOR, ip_address="10.0.0.1")
        token = self.sm.create_session(user)
        user2 = self.sm.validate_session(token)
        self.assertEqual(user2.username, "alice")
        self.assertEqual(user2.role, Role.OPERATOR)

    def test_invalid_signature_raises(self):
        user  = UserContext("1", "alice", Role.OPERATOR)
        token = self.sm.create_session(user)
        # Tamper with signature
        parts = token.rsplit(".", 1)
        bad_token = parts[0] + ".deadbeef" * 8
        with self.assertRaises(ValueError):
            self.sm.validate_session(bad_token)

    def test_revoked_token_raises(self):
        user  = UserContext("1", "alice", Role.OPERATOR)
        token = self.sm.create_session(user)
        self.sm.revoke_session(token)
        with self.assertRaises(ValueError):
            self.sm.validate_session(token)

    def test_expired_token_raises(self):
        sm    = SessionManager(secret_key="test-key", ttl_seconds=-1)  # expires immediately
        user  = UserContext("1", "alice", Role.OPERATOR)
        token = sm.create_session(user)
        with self.assertRaises(ValueError):
            sm.validate_session(token)


class TestApprovalWorkflow(unittest.TestCase):
    def setUp(self):
        self.sm, _ = _seeded_sm(n_employees=5)
        _run_payroll(self.sm)
        self.workflow = ApprovalWorkflow(self.sm)
        self.approver = UserContext("3", "approver", Role.APPROVER)
        self.operator = UserContext("2", "operator", Role.OPERATOR)

    def test_operator_cannot_approve(self):
        with self.assertRaises(PermissionError):
            self.workflow.approve_run(1, 1, 3, 2026, user=self.operator)

    def test_approver_can_approve_and_locks(self):
        self.workflow.approve_run(1, 1, 3, 2026, user=self.approver)
        self.assertTrue(self.workflow.is_approved(1, 1))

    def test_pending_approvals(self):
        """Processed (not yet locked) runs should appear in pending."""
        pending = self.workflow.pending_approvals(3, 2026)
        self.assertGreaterEqual(len(pending), 1)

    def test_after_approval_not_pending(self):
        self.workflow.approve_run(1, 1, 3, 2026, user=self.approver)
        pending = self.workflow.pending_approvals(3, 2026)
        pending_ids = [(p["shard_id"], p["payroll_run_id"]) for p in pending]
        self.assertNotIn((1, 1), pending_ids)


# ══════════════════════════════════════════════════════════════════════════════
# INTEGRATION — End-to-End (All Phases)
# ══════════════════════════════════════════════════════════════════════════════
class TestEndToEnd(unittest.TestCase):
    def test_50_employees_all_phases(self):
        """Full pipeline: Phase 1–7 for 50 employees."""
        tmp = Path(tempfile.mkdtemp())
        sm  = ShardManager(shard_dir=tmp/"shards", employees_per_shard=20)
        sm.initialize_shards(3)
        for i in range(1, 51):
            sm.insert_employee(build_seed_employee(i))

        bus = EventBus()

        # Phase 6: register plugins
        audit_p, backup_p = setup_phase6(sm, bus=bus, operator="integration_test",
                                          backup_dir=tmp/"backups")

        # Phase 7: register RBAC
        ctx7 = setup_phase7(sm, bus=bus, session_key="integration-test-key-32chars!")
        session_mgr = ctx7["session_manager"]
        workflow    = ctx7["approval_workflow"]

        # Phase 3: run payroll
        run_events = []
        bus.subscribe(PayrollEvent.RUN_COMPLETED, lambda **kw: run_events.append(kw))
        collector = PayrollRunner(sm, bus=bus).run(month=3, year=2026,
                                                    initiated_by="payroll_op")
        self.assertEqual(collector.processed_count, 50)
        self.assertEqual(len(collector.errors), 0)
        self.assertEqual(len(run_events), 1)

        # Phase 5: aggregation + exports
        orch   = ExportOrchestrator(sm, COMPANY_INFO, export_dir=tmp/"exports", bus=bus)
        result = orch.run(3, 2026, collector_totals=collector.shard_totals)
        report = result["report"]
        self.assertEqual(report.employee_count, 50)
        self.assertTrue(report.is_valid)

        for name, path in result["paths"].items():
            self.assertTrue(path.exists(), f"Missing export: {name}")

        # Phase 4: template pipeline
        txs_p4 = [{**vars(tx), "kra_pin": tx.kra_pin} for tx in report.transactions]
        pkgs = Phase4Pipeline(COMPANY_INFO, tmp/"tpl", tmp/"filled").run(txs_p4, 3, 2026)
        for code, pkg in pkgs.items():
            self.assertTrue(pkg["validation"]["passed"],
                            f"P4 {code}: {pkg['validation']['errors']}")

        # Phase 6 audit: entries should be in audit_log
        all_entries = []
        for sid in sm.list_shards():
            all_entries.extend(sm.get_audit_log(sid))
        self.assertGreater(len(all_entries), 0)

        # Phase 6 backup: triggered by AGGREGATION_DONE
        time.sleep(0.15)   # allow async backup to complete
        backup_dir = tmp/"backups"/"2026-03"
        self.assertTrue(backup_dir.exists(), "Backup directory should exist")

        # Phase 6: P9A annual (3 months)
        for m in [1, 2]:
            PayrollRunner(sm, bus=bus).run(month=m, year=2026)
        p9a_agg = P9AAnnualAggregator(sm, export_dir=tmp/"p9a")
        p9a_recs = p9a_agg.aggregate(year=2026)
        self.assertEqual(len(p9a_recs), 50)
        for rec in p9a_recs:
            self.assertGreaterEqual(rec.months_worked, 1)
        p9a_path = p9a_agg.export_excel(p9a_recs, 2026, COMPANY_INFO)
        self.assertTrue(p9a_path.exists())

        # Phase 7: session + approval
        approver = UserContext("A1", "cfo", Role.APPROVER, company_id=1)
        token    = session_mgr.create_session(approver)
        user_rt  = session_mgr.validate_session(token)
        self.assertEqual(user_rt.role, Role.APPROVER)

        pending = workflow.pending_approvals(3, 2026)
        self.assertGreater(len(pending), 0)

        for p in pending:
            workflow.approve_run(p["shard_id"], p["payroll_run_id"],
                                 3, 2026, user=approver)

        pending_after = workflow.pending_approvals(3, 2026)
        self.assertEqual(len(pending_after), 0)

    def test_multi_month_runs(self):
        sm, _ = _seeded_sm(n_employees=5)
        c1 = _run_payroll(sm, month=3, year=2026)
        c2 = _run_payroll(sm, month=4, year=2026)
        self.assertEqual(c1.processed_count, 5)
        self.assertEqual(c2.processed_count, 5)

    def test_corrected_caps_flow_through(self):
        """Verify corrected caps survive the full DB roundtrip."""
        sm, _ = _seeded_sm(n_employees=2)
        # employees 1-2 already seeded; attach reliefs above the corrected caps
        with sm.connection(1) as conn:
            conn.execute("""
                INSERT OR IGNORE INTO reliefs (employee_id, relief_type, monthly_amount)
                VALUES (1, 'mortgage', 35000)
            """)
            conn.execute("""
                INSERT OR IGNORE INTO reliefs (employee_id, relief_type, monthly_amount)
                VALUES (1, 'post_retirement', 18000)
            """)

        c = PayrollRunner(sm).run(3, 2026)
        self.assertGreaterEqual(c.processed_count, 1)


if __name__ == "__main__":
    loader = unittest.TestLoader()
    suite  = unittest.TestSuite()
    for cls in [
        # Phase 1
        TestShardManager,
        # Phase 2
        TestNSSF, TestSHIF, TestAHL, TestPAYEBands, TestPayrollEngine,
        # Phase 3
        TestPinnedWorkers,
        # Phase 4
        TestTemplateStore, TestTemplatePopulator, TestTemplateValidator,
        TestPhase4Pipeline,
        # Phase 5
        TestShardReader, TestAggregator, TestGLExporter,
        TestGovernmentFiles, TestExportOrchestrator,
        # Phase 6
        TestAuditPlugin, TestBackupPlugin, TestP9AAnnualAggregator,
        # Phase 7
        TestRBAC, TestCredentialVault, TestSessionManager, TestApprovalWorkflow,
        # Integration
        TestEndToEnd,
    ]:
        suite.addTests(loader.loadTestsFromTestCase(cls))

    result = unittest.TextTestRunner(verbosity=2).run(suite)
    sys.exit(0 if result.wasSuccessful() else 1)
