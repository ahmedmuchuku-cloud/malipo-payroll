"""
=============================================================================
KENYA PAYROLL SYSTEM — Phase 4: Authority Excel Template Automation
=============================================================================
Responsibilities:
  1. TemplateStore      — versioned template storage per month/year
  2. TemplateDownloader — abstract base + portal-specific implementations
                          (real Playwright/Selenium in production; stub here
                           so the system runs without portal connectivity)
  3. TemplatePopulator  — maps payroll_transactions → authority column layout
  4. TemplateValidator  — header, dtype, totals, completeness checks
  5. UploadPreparator   — formats filled templates for portal submission

Portal coverage: KRA PAYE (P10), NSSF Monthly Schedule, SHIF (SHA),
                 Affordable Housing Levy (AHL)

Design notes for future phases
───────────────────────────────
• Phase 6: validator writes a SHA-256 manifest so backups can detect tampering.
• Phase 7: TemplateDownloader.login() is the integration point for the
  encrypted credential vault.  Swap get_portal_credential() with vault lookup.
• Each populated template is written to /filled/{YYYY-MM}/ with a timestamp
  so the audit trail can reconstruct exactly what was submitted.
=============================================================================
"""

from __future__ import annotations

import csv
import hashlib
import json
import logging
import shutil
from abc import ABC, abstractmethod
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    TEMPLATE_DIR, FILLED_DIR, EXPORT_DIR, PORTALS, PortalConfig,
    get_portal_credential, STYLE, make_border, GL,
)

logger = logging.getLogger(__name__)


# ─── HELPERS ──────────────────────────────────────────────────────────────────
def _period_key(month: int, year: int) -> str:
    return f"{year}-{month:02d}"

def _money(v: float) -> float:
    return round(v, 2)

def _apply_header_style(ws, row: int, cols: int,
                         title: str = "", portal_color: str = "1a2e20") -> None:
    """Apply branded header styling to an openpyxl worksheet row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = PatternFill("solid", fgColor=portal_color.lstrip("#"))
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = make_border()

    if title:
        ws.cell(row=row, column=1).value = title

def _style_data_row(ws, row: int, cols: int, alt: bool = False) -> None:
    fill_hex = "101f15" if not alt else "0d1a10"
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = PatternFill("solid", fgColor=fill_hex)
        cell.font      = Font(name="Calibri", size=10, color="dff0e6")
        cell.border    = make_border()
        cell.alignment = Alignment(vertical="center")

def _style_total_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = PatternFill("solid", fgColor="0b1610")
        cell.font      = Font(bold=True, color="f59e0b", name="Calibri", size=10)
        cell.border    = make_border()
        cell.alignment = Alignment(horizontal="right" if col > 3 else "left",
                                   vertical="center")

def _freeze_and_autofit(ws, freeze_row: int = 4) -> None:
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


# ─── TEMPLATE STORE ───────────────────────────────────────────────────────────
class TemplateStore:
    """
    Manages versioned template files on disk.

    Layout:
      templates/{YYYY-MM}/kra_paye_template.xlsx
      templates/{YYYY-MM}/nssf_template.xlsx
      templates/{YYYY-MM}/shif_template.xlsx
      templates/{YYYY-MM}/ahl_template.xlsx

      filled/{YYYY-MM}/kra_paye_filled_{timestamp}.xlsx
      filled/{YYYY-MM}/nssf_filled_{timestamp}.xlsx
      ...

    A JSON manifest is written alongside filled templates for Phase 6 audit.
    """

    def __init__(self, template_dir: Path = TEMPLATE_DIR,
                 filled_dir: Path = FILLED_DIR):
        self.template_dir = Path(template_dir)
        self.filled_dir   = Path(filled_dir)

    def template_path(self, portal_code: str, month: int, year: int) -> Path:
        d = self.template_dir / _period_key(month, year)
        d.mkdir(parents=True, exist_ok=True)
        return d / f"{portal_code}_template.xlsx"

    def filled_path(self, portal_code: str, month: int, year: int) -> Path:
        d = self.filled_dir / _period_key(month, year)
        d.mkdir(parents=True, exist_ok=True)
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return d / f"{portal_code}_filled_{ts}.xlsx"

    def manifest_path(self, portal_code: str, month: int, year: int) -> Path:
        d = self.filled_dir / _period_key(month, year)
        d.mkdir(parents=True, exist_ok=True)
        return d / f"{portal_code}_manifest.json"

    def save_manifest(self, portal_code: str, month: int, year: int,
                      filled_path: Path, row_count: int,
                      checksum: str, totals: dict) -> None:
        """Write a manifest JSON for Phase 6 audit / tamper detection."""
        manifest = {
            "portal":       portal_code,
            "period":       _period_key(month, year),
            "generated_at": datetime.utcnow().isoformat(),
            "file":         str(filled_path.name),
            "sha256":       checksum,
            "row_count":    row_count,
            "totals":       totals,
        }
        path = self.manifest_path(portal_code, month, year)
        with open(path, "w") as f:
            json.dump(manifest, f, indent=2)
        logger.info(f"Manifest saved: {path}")

    def list_filled(self, portal_code: str, month: int, year: int) -> list[Path]:
        d = self.filled_dir / _period_key(month, year)
        if not d.exists():
            return []
        return sorted(d.glob(f"{portal_code}_filled_*.xlsx"))

    @staticmethod
    def sha256(path: Path) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()


# ─── BASE DOWNLOADER ─────────────────────────────────────────────────────────
class BaseTemplateDownloader(ABC):
    """
    Abstract template downloader.

    In production, subclass this with Playwright/Selenium to:
      1. Navigate to portal login URL
      2. Submit credentials (from Phase 7 vault)
      3. Handle OTP / 2FA
      4. Download the monthly Excel template
      5. Save to TemplateStore

    The stub implementations below generate empty-but-valid templates
    so the rest of the pipeline works without portal access.
    This is also how automated testing works.
    """

    def __init__(self, portal: PortalConfig, store: TemplateStore):
        self.portal = portal
        self.store  = store

    def download(self, month: int, year: int,
                 company_pin: str = "") -> Path:
        """
        Downloads (or generates stub) template for the given period.
        Returns path to saved template file.
        """
        dest = self.store.template_path(self.portal.code, month, year)
        if dest.exists():
            logger.info(f"Template already cached: {dest}")
            return dest

        logger.info(f"Downloading {self.portal.name} template for {_period_key(month, year)}")
        wb = self._fetch_or_stub(month, year, company_pin)
        wb.save(str(dest))
        logger.info(f"Template saved: {dest}")
        return dest

    def login(self, username: str, password: str) -> bool:
        """
        Phase 7 integration point: replace body with authenticated session.
        Returns True if login succeeded.
        """
        # Stub: always "succeeds" — Phase 7 replaces with Playwright session
        logger.debug(f"[STUB] login to {self.portal.login_url} as {username}")
        return True

    @abstractmethod
    def _fetch_or_stub(self, month: int, year: int, company_pin: str) -> openpyxl.Workbook:
        """Return a Workbook representing the authority template."""
        ...


# ─── PORTAL-SPECIFIC DOWNLOADERS ─────────────────────────────────────────────
class KRAPayeDownloader(BaseTemplateDownloader):
    """KRA iTax PAYE monthly return template (P10 format)."""

    def _fetch_or_stub(self, month: int, year: int, company_pin: str) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P10 Return"
        ws.append(["KRA P10 PAYE MONTHLY RETURN", "", "", "", "", "", "", ""])
        ws.append([f"Period: {_period_key(month, year)}", "", "", "",
                   "Employer PIN:", company_pin, "", ""])
        ws.append([])
        ws.append([
            "EMPLOYEE_NO", "KRA_PIN", "EMPLOYEE_NAME", "BASIC_PAY",
            "TAXABLE_ALLOWANCES", "NON_TAXABLE_ALLOWANCES", "BENEFITS_IN_KIND",
            "NSSF_EMPLOYEE", "SHIF", "AHL_EMPLOYEE",
            "TAXABLE_INCOME", "GROSS_PAYE", "PERSONAL_RELIEF",
            "INSURANCE_RELIEF", "MORTGAGE_RELIEF", "AHL_RELIEF",
            "OTHER_RELIEF", "TOTAL_RELIEF", "NET_PAYE",
        ])
        return wb


class NSSFDownloader(BaseTemplateDownloader):
    """NSSF monthly contribution schedule template."""

    def _fetch_or_stub(self, month: int, year: int, company_pin: str) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NSSF Schedule"
        ws.append(["NSSF MONTHLY CONTRIBUTION SCHEDULE", "", "", "", "", "", ""])
        ws.append([f"Period: {_period_key(month, year)}", "", "", "Employer No.:", "", "", ""])
        ws.append([])
        ws.append([
            "EMPLOYEE_NO", "NSSF_MEMBER_NO", "EMPLOYEE_NAME",
            "GROSS_SALARY", "TIER1_EMPLOYEE", "TIER2_EMPLOYEE",
            "TOTAL_EMPLOYEE", "TIER1_EMPLOYER", "TIER2_EMPLOYER",
            "TOTAL_EMPLOYER", "GRAND_TOTAL",
        ])
        return wb


class SHIFDownloader(BaseTemplateDownloader):
    """SHA SHIF monthly contribution template."""

    def _fetch_or_stub(self, month: int, year: int, company_pin: str) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SHIF Schedule"
        ws.append(["SHIF MONTHLY CONTRIBUTION — SHA PORTAL", "", "", "", ""])
        ws.append([f"Period: {_period_key(month, year)}", "", "Rate: 2.75%", "", ""])
        ws.append([])
        ws.append([
            "EMPLOYEE_NO", "SHA_MEMBER_NO", "EMPLOYEE_NAME",
            "GROSS_SALARY", "SHIF_CONTRIBUTION", "STATUS",
        ])
        return wb


class AHLDownloader(BaseTemplateDownloader):
    """KRA iTax Affordable Housing Levy return template."""

    def _fetch_or_stub(self, month: int, year: int, company_pin: str) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AHL Return"
        ws.append(["AFFORDABLE HOUSING LEVY MONTHLY RETURN", "", "", "", ""])
        ws.append([f"Period: {_period_key(month, year)}", "", "Employer PIN:", company_pin, ""])
        ws.append([])
        ws.append([
            "EMPLOYEE_NO", "KRA_PIN", "EMPLOYEE_NAME",
            "GROSS_SALARY", "AHL_EMPLOYEE_1.5PCT", "AHL_EMPLOYER_1.5PCT",
            "TOTAL_AHL", "AHL_RELIEF",
        ])
        return wb


# Registry — extend for future portals without touching caller code
DOWNLOADERS: dict[str, type[BaseTemplateDownloader]] = {
    "kra":  KRAPayeDownloader,
    "nssf": NSSFDownloader,
    "shif": SHIFDownloader,
    "ahl":  AHLDownloader,
}

def get_downloader(portal_code: str, store: TemplateStore) -> BaseTemplateDownloader:
    cls = DOWNLOADERS.get(portal_code)
    if not cls:
        raise ValueError(f"No downloader for portal '{portal_code}'")
    return cls(PORTALS[portal_code], store)


# ─── VALIDATION RESULT ────────────────────────────────────────────────────────
@dataclass
class ValidationResult:
    portal_code: str
    month:       int
    year:        int
    passed:      bool = True
    errors:      list[str] = field(default_factory=list)
    warnings:    list[str] = field(default_factory=list)
    row_count:   int  = 0
    totals:      dict = field(default_factory=dict)

    def fail(self, msg: str):
        self.passed = False
        self.errors.append(msg)

    def warn(self, msg: str):
        self.warnings.append(msg)


# ─── TEMPLATE POPULATOR ───────────────────────────────────────────────────────
class TemplatePopulator:
    """
    Maps payroll transaction data (from Phase 5 Aggregator or shard reads)
    into the authority template column layout, applies styling, and saves.
    """

    def __init__(self, store: TemplateStore):
        self.store = store

    # ── KRA PAYE (P10) ───────────────────────────────────────────────────────
    def populate_kra_paye(self, transactions: list[dict],
                          month: int, year: int,
                          company_info: dict) -> Path:
        """
        Populates the KRA PAYE P10 template.
        transactions: list of payroll_transaction dicts from any shard.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P10 Return"

        period = _period_key(month, year)
        color  = "166534"   # KRA green

        # Title block
        ws.merge_cells("A1:S1")
        ws["A1"] = f"KRA P10 PAYE MONTHLY RETURN — {period}"
        ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
        ws["A1"].fill      = PatternFill("solid", fgColor=color)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = f"Employer: {company_info.get('name', '')}"
        ws["D2"] = f"PIN: {company_info.get('pin', '')}"
        ws["G2"] = f"NSSF No.: {company_info.get('nssf_no', '')}"
        ws["J2"] = f"Period: {period}"
        ws["A2"].font = Font(bold=True, size=10, name="Calibri")

        # Column headers (row 4)
        HEADERS = [
            ("EMP_NO",           8),
            ("KRA_PIN",         14),
            ("EMPLOYEE_NAME",   28),
            ("BASIC_PAY",       14),
            ("TAXABLE_ALLOW",   14),
            ("NON_TAX_ALLOW",   14),
            ("BIK",             10),
            ("NSSF_EMP",        12),
            ("SHIF",            12),
            ("AHL_EMP",         10),
            ("TAXABLE_INCOME",  16),
            ("GROSS_PAYE",      14),
            ("PERSONAL_REL",    14),
            ("INSURANCE_REL",   14),
            ("MORTGAGE_REL",    14),
            ("AHL_RELIEF",      12),
            ("OTHER_REL",       12),
            ("TOTAL_RELIEF",    14),
            ("NET_PAYE",        14),
        ]
        for col, (hdr, width) in enumerate(HEADERS, 1):
            cell = ws.cell(row=4, column=col, value=hdr)
            cell.fill      = PatternFill("solid", fgColor=color)
            cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = make_border()
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[4].height = 30

        # Data rows
        total_cols = {
            "BASIC_PAY": 0.0, "TAXABLE_ALLOW": 0.0, "BIK": 0.0,
            "NSSF_EMP": 0.0,  "SHIF": 0.0,           "AHL_EMP": 0.0,
            "TAXABLE_INCOME": 0.0, "GROSS_PAYE": 0.0,
            "TOTAL_RELIEF": 0.0,   "NET_PAYE": 0.0,
        }
        money_cols = set(range(4, 20))   # columns with KES values

        for i, tx in enumerate(transactions):
            r     = 5 + i
            alt   = (i % 2 == 1)
            fill  = PatternFill("solid", fgColor="101f15" if not alt else "0d1a10")

            values = [
                tx.get("employee_id", ""),
                tx.get("kra_pin", ""),
                tx.get("full_name", ""),
                _money(tx.get("base_salary", 0)),
                _money(tx.get("taxable_allowances", 0)),
                _money(tx.get("non_taxable_allowances", 0)),
                _money(tx.get("benefits_in_kind", 0)),
                _money(tx.get("nssf_employee", 0)),
                _money(tx.get("shif", 0)),
                _money(tx.get("housing_levy_employee", 0)),
                _money(tx.get("taxable_income", 0)),
                _money(tx.get("gross_paye", 0)),
                _money(tx.get("personal_relief", 0)),
                _money(tx.get("insurance_relief", 0)),
                _money(tx.get("mortgage_relief", 0)),
                _money(tx.get("ahl_relief", 0)),
                _money(tx.get("post_retirement_relief", 0)),
                _money(tx.get("total_relief", 0)),
                _money(tx.get("paye", 0)),
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill      = fill
                cell.font      = Font(size=9, color="dff0e6", name="Calibri")
                cell.border    = make_border()
                if col in money_cols and isinstance(val, (int, float)):
                    cell.number_format = "#,##0.00"
                    cell.alignment     = Alignment(horizontal="right")

            # Running totals
            total_cols["BASIC_PAY"]      += tx.get("base_salary", 0)
            total_cols["NSSF_EMP"]       += tx.get("nssf_employee", 0)
            total_cols["SHIF"]           += tx.get("shif", 0)
            total_cols["AHL_EMP"]        += tx.get("housing_levy_employee", 0)
            total_cols["TAXABLE_INCOME"] += tx.get("taxable_income", 0)
            total_cols["GROSS_PAYE"]     += tx.get("gross_paye", 0)
            total_cols["TOTAL_RELIEF"]   += tx.get("total_relief", 0)
            total_cols["NET_PAYE"]       += tx.get("paye", 0)

        # Totals row
        total_row = 5 + len(transactions)
        ws.cell(row=total_row, column=1, value="TOTALS")
        totals_map = {4: total_cols["BASIC_PAY"],
                      8: total_cols["NSSF_EMP"],
                      9: total_cols["SHIF"],
                      10: total_cols["AHL_EMP"],
                      11: total_cols["TAXABLE_INCOME"],
                      12: total_cols["GROSS_PAYE"],
                      18: total_cols["TOTAL_RELIEF"],
                      19: total_cols["NET_PAYE"]}
        for col in range(1, 20):
            cell = ws.cell(row=total_row, column=col)
            if col in totals_map:
                cell.value         = _money(totals_map[col])
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right")
            cell.fill   = PatternFill("solid", fgColor="0b1610")
            cell.font   = Font(bold=True, color="f59e0b", size=9, name="Calibri")
            cell.border = make_border()

        ws.freeze_panes = "A5"

        out_path = self.store.filled_path("kra", month, year)
        wb.save(str(out_path))

        checksum = self.store.sha256(out_path)
        self.store.save_manifest("kra", month, year, out_path,
                                 len(transactions), checksum, total_cols)
        logger.info(f"KRA PAYE template populated: {out_path} ({len(transactions)} rows)")
        return out_path

    # ── NSSF ─────────────────────────────────────────────────────────────────
    def populate_nssf(self, transactions: list[dict],
                      month: int, year: int,
                      company_info: dict) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NSSF Schedule"
        period   = _period_key(month, year)
        color    = "1e3a8a"   # NSSF blue

        ws.merge_cells("A1:K1")
        ws["A1"] = f"NSSF MONTHLY CONTRIBUTION SCHEDULE — {period}"
        ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        ws["A1"].fill      = PatternFill("solid", fgColor=color)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = f"Employer: {company_info.get('name', '')}"
        ws["D2"] = f"NSSF No.: {company_info.get('nssf_no', '')}"

        HEADERS = [
            ("EMP_NO", 8), ("NSSF_MEMBER_NO", 16), ("EMPLOYEE_NAME", 28),
            ("GROSS_SALARY", 16), ("TIER1_EMPLOYEE", 14), ("TIER2_EMPLOYEE", 14),
            ("TOTAL_EMPLOYEE", 14), ("TIER1_EMPLOYER", 14), ("TIER2_EMPLOYER", 14),
            ("TOTAL_EMPLOYER", 14), ("GRAND_TOTAL", 14),
        ]
        for col, (hdr, width) in enumerate(HEADERS, 1):
            cell = ws.cell(row=4, column=col, value=hdr)
            cell.fill      = PatternFill("solid", fgColor=color)
            cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = make_border()
            ws.column_dimensions[get_column_letter(col)].width = width

        t_emp = 0.0; t_emplr = 0.0; t_gross = 0.0

        for i, tx in enumerate(transactions):
            r    = 5 + i
            alt  = (i % 2 == 1)
            fill = PatternFill("solid", fgColor="101f15" if not alt else "0d1a10")

            nssf_emp   = _money(tx.get("nssf_employee", 0))
            nssf_emplr = _money(tx.get("nssf_employer", 0))
            gross      = _money(tx.get("gross_salary", 0))
            t1e        = _money(tx.get("nssf_tier1_employee", 0))
            t2e        = _money(tx.get("nssf_tier2_employee", 0))
            t1r        = _money(tx.get("nssf_tier1_employer", 0))
            t2r        = _money(tx.get("nssf_tier2_employer", 0))

            values = [
                tx.get("employee_id", ""), tx.get("nssf_number", ""),
                tx.get("full_name", ""), gross,
                t1e, t2e, nssf_emp, t1r, t2r, nssf_emplr,
                _money(nssf_emp + nssf_emplr),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill   = fill
                cell.font   = Font(size=9, color="dff0e6", name="Calibri")
                cell.border = make_border()
                if col >= 4 and isinstance(val, float):
                    cell.number_format = "#,##0.00"
                    cell.alignment     = Alignment(horizontal="right")

            t_gross += gross; t_emp += nssf_emp; t_emplr += nssf_emplr

        total_row = 5 + len(transactions)
        totals = [
            "TOTALS", "", "", _money(t_gross),
            "", "", _money(t_emp), "", "", _money(t_emplr),
            _money(t_emp + t_emplr),
        ]
        for col, val in enumerate(totals, 1):
            cell = ws.cell(row=total_row, column=col, value=val)
            cell.fill   = PatternFill("solid", fgColor="0b1610")
            cell.font   = Font(bold=True, color="f59e0b", size=9, name="Calibri")
            cell.border = make_border()
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right")

        ws.freeze_panes = "A5"
        out_path = self.store.filled_path("nssf", month, year)
        wb.save(str(out_path))

        checksum = self.store.sha256(out_path)
        self.store.save_manifest("nssf", month, year, out_path, len(transactions),
                                 checksum, {"nssf_employee": t_emp,
                                            "nssf_employer": t_emplr,
                                            "gross": t_gross})
        logger.info(f"NSSF template populated: {out_path}")
        return out_path

    # ── SHIF ─────────────────────────────────────────────────────────────────
    def populate_shif(self, transactions: list[dict],
                      month: int, year: int,
                      company_info: dict) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SHIF Schedule"
        period   = _period_key(month, year)
        color    = "854d0e"

        ws.merge_cells("A1:F1")
        ws["A1"] = f"SHIF MONTHLY CONTRIBUTION — SHA PORTAL — {period}"
        ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        ws["A1"].fill      = PatternFill("solid", fgColor=color)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"] = f"Employer: {company_info.get('name', '')} | SHA No.: {company_info.get('sha_no', '')} | Rate: 2.75%"

        HEADERS = [
            ("EMP_NO", 8), ("SHA_MEMBER_NO", 18), ("EMPLOYEE_NAME", 30),
            ("GROSS_SALARY", 18), ("SHIF_2.75PCT", 16), ("STATUS", 12),
        ]
        for col, (hdr, width) in enumerate(HEADERS, 1):
            cell = ws.cell(row=4, column=col, value=hdr)
            cell.fill      = PatternFill("solid", fgColor=color)
            cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="center")
            cell.border    = make_border()
            ws.column_dimensions[get_column_letter(col)].width = width

        t_shif = 0.0
        for i, tx in enumerate(transactions):
            r    = 5 + i
            alt  = (i % 2 == 1)
            fill = PatternFill("solid", fgColor="101f15" if not alt else "0d1a10")
            shif = _money(tx.get("shif", 0))
            values = [
                tx.get("employee_id", ""),
                tx.get("sha_number", f"SHA/{tx.get('employee_id', ''):08}"),
                tx.get("full_name", ""),
                _money(tx.get("gross_salary", 0)),
                shif,
                "Active",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill   = fill
                cell.font   = Font(size=9, color="dff0e6", name="Calibri")
                cell.border = make_border()
                if col in (4, 5) and isinstance(val, float):
                    cell.number_format = "#,##0.00"
                    cell.alignment     = Alignment(horizontal="right")
            t_shif += shif

        total_row = 5 + len(transactions)
        for col in range(1, 7):
            cell = ws.cell(row=total_row, column=col)
            cell.fill   = PatternFill("solid", fgColor="0b1610")
            cell.font   = Font(bold=True, color="f59e0b", size=9, name="Calibri")
            cell.border = make_border()
        ws.cell(row=total_row, column=1).value = "TOTALS"
        ws.cell(row=total_row, column=5).value = _money(t_shif)
        ws.cell(row=total_row, column=5).number_format = "#,##0.00"

        ws.freeze_panes = "A5"
        out_path = self.store.filled_path("shif", month, year)
        wb.save(str(out_path))
        checksum = self.store.sha256(out_path)
        self.store.save_manifest("shif", month, year, out_path, len(transactions),
                                 checksum, {"shif_total": t_shif})
        logger.info(f"SHIF template populated: {out_path}")
        return out_path

    # ── AHL ──────────────────────────────────────────────────────────────────
    def populate_ahl(self, transactions: list[dict],
                     month: int, year: int,
                     company_info: dict) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AHL Return"
        period   = _period_key(month, year)
        color    = "4c1d95"

        ws.merge_cells("A1:H1")
        ws["A1"] = f"AFFORDABLE HOUSING LEVY MONTHLY RETURN — {period}"
        ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        ws["A1"].fill      = PatternFill("solid", fgColor=color)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"] = f"Employer: {company_info.get('name', '')} | PIN: {company_info.get('pin', '')}"

        HEADERS = [
            ("EMP_NO", 8), ("KRA_PIN", 14), ("EMPLOYEE_NAME", 28),
            ("GROSS_SALARY", 16), ("AHL_EMPLOYEE_1.5%", 18),
            ("AHL_EMPLOYER_1.5%", 18), ("TOTAL_AHL", 14), ("AHL_RELIEF", 14),
        ]
        for col, (hdr, width) in enumerate(HEADERS, 1):
            cell = ws.cell(row=4, column=col, value=hdr)
            cell.fill      = PatternFill("solid", fgColor=color)
            cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = make_border()
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[4].height = 30

        t_emp = 0.0; t_emplr = 0.0; t_relief = 0.0
        for i, tx in enumerate(transactions):
            r    = 5 + i
            alt  = (i % 2 == 1)
            fill = PatternFill("solid", fgColor="101f15" if not alt else "0d1a10")
            ahl_e = _money(tx.get("housing_levy_employee", 0))
            ahl_r = _money(tx.get("housing_levy_employer", 0))
            rel   = _money(tx.get("ahl_relief", 0))
            values = [
                tx.get("employee_id", ""), tx.get("kra_pin", ""),
                tx.get("full_name", ""),
                _money(tx.get("gross_salary", 0)),
                ahl_e, ahl_r, _money(ahl_e + ahl_r), rel,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill   = fill
                cell.font   = Font(size=9, color="dff0e6", name="Calibri")
                cell.border = make_border()
                if col >= 4 and isinstance(val, float):
                    cell.number_format = "#,##0.00"
                    cell.alignment     = Alignment(horizontal="right")
            t_emp += ahl_e; t_emplr += ahl_r; t_relief += rel

        total_row = 5 + len(transactions)
        for col in range(1, 9):
            cell = ws.cell(row=total_row, column=col)
            cell.fill   = PatternFill("solid", fgColor="0b1610")
            cell.font   = Font(bold=True, color="f59e0b", size=9, name="Calibri")
            cell.border = make_border()
        ws.cell(row=total_row, column=1).value = "TOTALS"
        for col, val in [(5, t_emp), (6, t_emplr),
                          (7, _money(t_emp + t_emplr)), (8, t_relief)]:
            cell = ws.cell(row=total_row, column=col, value=_money(val))
            cell.number_format = "#,##0.00"
            cell.alignment     = Alignment(horizontal="right")

        ws.freeze_panes = "A5"
        out_path = self.store.filled_path("ahl", month, year)
        wb.save(str(out_path))
        checksum = self.store.sha256(out_path)
        self.store.save_manifest("ahl", month, year, out_path, len(transactions),
                                 checksum, {"ahl_employee": t_emp,
                                            "ahl_employer": t_emplr})
        logger.info(f"AHL template populated: {out_path}")
        return out_path


# ─── TEMPLATE VALIDATOR ───────────────────────────────────────────────────────
class TemplateValidator:
    """
    Validates a populated authority template before upload.
    Checks headers, data types, missing values, and total reconciliation.
    """

    EXPECTED_HEADERS: dict[str, list[str]] = {
        "kra":  ["EMP_NO", "KRA_PIN", "EMPLOYEE_NAME", "BASIC_PAY"],
        "nssf": ["EMP_NO", "NSSF_MEMBER_NO", "EMPLOYEE_NAME", "GROSS_SALARY"],
        "shif": ["EMP_NO", "SHA_MEMBER_NO",  "EMPLOYEE_NAME", "GROSS_SALARY"],
        "ahl":  ["EMP_NO", "KRA_PIN",         "EMPLOYEE_NAME", "GROSS_SALARY"],
    }

    # Column indices (1-based) that must contain numeric values
    NUMERIC_COLS: dict[str, list[int]] = {
        "kra":  list(range(4, 20)),
        "nssf": list(range(4, 12)),
        "shif": [4, 5],
        "ahl":  list(range(4, 9)),
    }

    def validate(self, path: Path, portal_code: str,
                 expected_row_count: int,
                 expected_totals: dict | None = None) -> ValidationResult:
        vr = ValidationResult(portal_code=portal_code,
                              month=0, year=0)  # caller can set period

        if not path.exists():
            vr.fail(f"File not found: {path}")
            return vr

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception as e:
            vr.fail(f"Cannot open workbook: {e}")
            return vr

        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        # Find header row (row 4 = index 3 in 0-based)
        if len(all_rows) < 5:
            vr.fail("Workbook has fewer than 5 rows — likely empty")
            return vr

        header_row = [str(c).strip() if c else "" for c in all_rows[3]]
        expected   = self.EXPECTED_HEADERS.get(portal_code, [])
        for h in expected:
            if h not in header_row:
                vr.fail(f"Missing expected header: '{h}'")

        # Data rows (row 5 onward, skip totals row at end)
        data_rows = all_rows[4:-1]   # exclude totals row
        vr.row_count = len(data_rows)

        if vr.row_count != expected_row_count:
            vr.fail(
                f"Row count mismatch: expected {expected_row_count}, "
                f"found {vr.row_count}"
            )

        # Numeric column check
        num_cols = self.NUMERIC_COLS.get(portal_code, [])
        for row_idx, row in enumerate(data_rows, start=5):
            for col_idx in num_cols:
                val = row[col_idx - 1] if col_idx <= len(row) else None
                if val is not None and val != "" and not isinstance(val, (int, float)):
                    vr.warn(
                        f"Row {row_idx}, col {col_idx}: "
                        f"expected numeric, got '{val}'"
                    )

        # Missing KRA PIN check (PAYE and AHL)
        if portal_code in ("kra", "ahl"):
            pin_col = 1   # 0-based in data_rows = column 2 (KRA_PIN)
            for row_idx, row in enumerate(data_rows, start=5):
                pin = row[pin_col] if len(row) > pin_col else None
                if not pin:
                    vr.warn(f"Row {row_idx}: missing KRA PIN")

        # Totals reconciliation
        if expected_totals:
            totals_row = all_rows[-1]
            self._reconcile_totals(vr, portal_code, totals_row, expected_totals)

        wb.close()
        vr.passed = len(vr.errors) == 0
        return vr

    def _reconcile_totals(self, vr: ValidationResult, portal_code: str,
                           totals_row: tuple, expected: dict) -> None:
        tol = 0.05   # KES 0.05 rounding tolerance

        mappings = {
            "kra":  {8: "nssf_employee", 9: "shif", 10: "ahl_employee",
                     12: "gross_paye", 19: "net_paye"},
            "nssf": {7: "nssf_employee", 10: "nssf_employer", 11: "grand_total"},
            "shif": {5: "shif_total"},
            "ahl":  {5: "ahl_employee", 6: "ahl_employer", 7: "ahl_total"},
        }

        for col_idx, key in mappings.get(portal_code, {}).items():
            if key not in expected:
                continue
            actual = totals_row[col_idx - 1] if col_idx <= len(totals_row) else None
            if actual is None:
                vr.warn(f"Totals row missing col {col_idx} ({key})")
                continue
            exp_v = expected[key]
            if abs(float(actual) - float(exp_v)) > tol:
                vr.fail(
                    f"Total mismatch for {key}: "
                    f"template={actual:.2f}, expected={exp_v:.2f}"
                )


# ─── UPLOAD PREPARATOR ────────────────────────────────────────────────────────
class UploadPreparator:
    """
    Prepares a filled template for portal upload.
    Produces CSV variants where portals require them, and a
    machine-readable upload manifest referencing the filled .xlsx.

    Phase 7 extension point: add digital signature here.
    """

    def __init__(self, store: TemplateStore):
        self.store = store

    def prepare(self, filled_path: Path, portal_code: str,
                 month: int, year: int) -> dict:
        """
        Returns an upload_package dict:
          {
            "portal":       "kra",
            "xlsx_path":    Path(...),
            "csv_path":     Path(...),    # if portal accepts CSV
            "manifest":     {...},
            "ready":        True/False,
            "upload_url":   "https://..."
          }
        """
        manifest_path = self.store.manifest_path(portal_code, month, year)
        manifest      = {}
        if manifest_path.exists():
            with open(manifest_path) as f:
                manifest = json.load(f)

        csv_path = self._xlsx_to_csv(filled_path, portal_code, month, year)

        portal = PORTALS.get(portal_code)
        return {
            "portal":      portal_code,
            "portal_name": portal.name if portal else portal_code,
            "period":      _period_key(month, year),
            "xlsx_path":   filled_path,
            "csv_path":    csv_path,
            "manifest":    manifest,
            "sha256":      manifest.get("sha256", ""),
            "row_count":   manifest.get("row_count", 0),
            "totals":      manifest.get("totals", {}),
            "upload_url":  portal.template_url if portal else "",
            "paybill":     portal.paybill if portal else "",
            "ready":       filled_path.exists(),
        }

    def _xlsx_to_csv(self, xlsx_path: Path, portal_code: str,
                     month: int, year: int) -> Path | None:
        """Convert filled xlsx to CSV (some portals need CSV upload)."""
        try:
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
            ws = wb.active
            d  = self.store.filled_dir / _period_key(month, year)
            d.mkdir(parents=True, exist_ok=True)
            csv_path = d / f"{portal_code}_upload.csv"
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(min_row=4, values_only=True):
                    writer.writerow(row)
            wb.close()
            return csv_path
        except Exception as e:
            logger.warning(f"CSV export failed for {portal_code}: {e}")
            return None


# ─── FACADE ───────────────────────────────────────────────────────────────────
class Phase4Pipeline:
    """
    Single entry point for the full Phase 4 workflow:
      download → populate → validate → prepare upload

    Usage:
        p4 = Phase4Pipeline(company_info={...})
        packages = p4.run(transactions, month=3, year=2026)
    """

    def __init__(self,
                 company_info:  dict,
                 template_dir:  Path = TEMPLATE_DIR,
                 filled_dir:    Path = FILLED_DIR):
        self.company_info = company_info
        self.store        = TemplateStore(template_dir, filled_dir)
        self.populator    = TemplatePopulator(self.store)
        self.validator    = TemplateValidator()
        self.preparator   = UploadPreparator(self.store)

    def run(self, transactions: list[dict],
            month: int, year: int,
            portals: list[str] = ("kra", "nssf", "shif", "ahl"),
            ) -> dict[str, dict]:
        """
        Run the full pipeline for all specified portals.
        Returns {portal_code: upload_package}.
        """
        # Augment transactions with employee lookup data if present
        results: dict[str, dict] = {}

        populate_fns = {
            "kra":  self.populator.populate_kra_paye,
            "nssf": self.populator.populate_nssf,
            "shif": self.populator.populate_shif,
            "ahl":  self.populator.populate_ahl,
        }

        for portal_code in portals:
            logger.info(f"Phase 4: processing portal '{portal_code}'")

            # 1. Download / ensure template exists
            downloader = get_downloader(portal_code, self.store)
            downloader.download(month, year, self.company_info.get("pin", ""))

            # 2. Populate
            fn = populate_fns.get(portal_code)
            if not fn:
                logger.warning(f"No populator for {portal_code}")
                continue
            filled_path = fn(transactions, month, year, self.company_info)

            # 3. Validate
            vr = self.validator.validate(
                filled_path, portal_code,
                expected_row_count=len(transactions),
            )
            if not vr.passed:
                logger.error(f"Validation FAILED for {portal_code}: {vr.errors}")
            else:
                logger.info(f"Validation PASSED for {portal_code} "
                            f"({len(vr.warnings)} warnings)")

            # 4. Prepare upload package
            package = self.preparator.prepare(filled_path, portal_code, month, year)
            package["validation"] = {
                "passed":   vr.passed,
                "errors":   vr.errors,
                "warnings": vr.warnings,
            }
            results[portal_code] = package

        return results
