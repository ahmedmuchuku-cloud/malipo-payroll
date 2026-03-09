"""
=============================================================================
KENYA PAYROLL SYSTEM — Phase 6: Audit & Backup
=============================================================================
Provides three production-grade components:

  1. AuditPlugin       — Writes structured audit entries to each shard's
                         audit_log table on every transaction, error, and
                         template event. Subscribes to the EventBus; zero
                         coupling to core payroll code.

  2. BackupPlugin      — Triggered on AGGREGATION_DONE; snapshots every
                         live shard using SQLite's native backup API
                         (safe while DB is live). Writes a SHA-256 manifest,
                         purges backups older than BACKUP_RETENTION_DAYS,
                         and logs a summary.

  3. P9AAnnualAggregator — Reads all 12 monthly payroll runs for a given
                           tax year and produces per-employee annual P9A
                           certificates (Excel), the legally required form
                           issued by employers before end of February.

Design principles:
  • AuditPlugin and BackupPlugin are pure EventBus subscribers — they
    never call payroll engine code directly.
  • A broken plugin NEVER crashes a payroll run (EventBus swallows
    exceptions from handlers).
  • All file paths are deterministic: /backups/{YYYY-MM}/{shard_id}_{ts}.db
  • SHA-256 checksums on backup files enable tamper detection on restore.
=============================================================================
"""
from __future__ import annotations

import hashlib
import json
import logging
import shutil
import sqlite3
import threading
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import BACKUP_DIR, BACKUP_RETENTION_DAYS, EXPORT_DIR, STYLE, make_border
from events import EventBus, PayrollEvent, PayrollPlugin, get_bus
from shard_manager import ShardManager

logger = logging.getLogger(__name__)


# ─── HELPERS ─────────────────────────────────────────────────────────────────

def _period(month: int, year: int) -> str:
    return f"{year}-{month:02d}"

def _now_ts() -> str:
    return datetime.utcnow().strftime("%Y%m%d_%H%M%S")

def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ─── AUDIT PLUGIN ─────────────────────────────────────────────────────────────

class AuditPlugin(PayrollPlugin):
    """
    Writes structured audit entries to each shard's audit_log table.

    Hooks:
      TRANSACTION_SAVED  → one CALC entry per employee per run
      ERROR_OCCURRED     → one ERROR entry
      TEMPLATE_POPULATED → one TEMPLATE entry per portal
      RUN_COMPLETED      → one RUN_SUMMARY entry to all shards

    All writes are thread-safe (one lock per shard, inherited from ShardManager).
    The operator field carries the initiated_by value from the PayrollRunner.
    """
    name = "AuditPlugin"

    def __init__(self, shard_manager: ShardManager, operator: str = "system"):
        self.sm       = shard_manager
        self.operator = operator
        self._run_operator: dict[str, str] = {}   # run_id → operator

    # ── Registration ──────────────────────────────────────────────────────────
    def register(self, bus: EventBus | None = None) -> None:
        b = bus or get_bus()
        b.subscribe(PayrollEvent.RUN_STARTED,        self.on_run_started)
        b.subscribe(PayrollEvent.TRANSACTION_SAVED,  self.on_transaction_saved)
        b.subscribe(PayrollEvent.ERROR_OCCURRED,     self.on_error)
        b.subscribe(PayrollEvent.TEMPLATE_POPULATED, self.on_template_populated)
        b.subscribe(PayrollEvent.RUN_COMPLETED,      self.on_run_completed)
        b.subscribe(PayrollEvent.EXPORT_CREATED,     self.on_export_created)
        logger.info(f"Plugin '{self.name}' registered")

    # ── Handlers ──────────────────────────────────────────────────────────────
    def on_run_started(self, event, **kw):
        # Store the initiating operator from RUN_STARTED payload if provided
        initiated_by = kw.get("initiated_by", self.operator)
        run_map = kw.get("run_map", {})
        for run_id in run_map.values():
            self._run_operator[str(run_id)] = initiated_by

    def on_transaction_saved(self, event, **kw):
        employee_id = kw.get("employee_id", 0)
        shard_id    = kw.get("shard_id", 0)
        net_pay     = kw.get("net_pay", 0.0)
        run_id      = kw.get("payroll_run_id")

        if not shard_id:
            return

        self._write(shard_id, dict(
            payroll_run_id = run_id,
            employee_id    = employee_id,
            event_type     = "CALC",
            rule_applied   = "PAYROLL_ENGINE_2026",
            field_name     = "net_pay",
            before_value   = None,
            after_value    = str(round(net_pay, 2)),
            notes          = f"Payroll computed — net_pay={net_pay:.2f}",
            operator       = self._run_operator.get(str(run_id), self.operator),
        ))

    def on_error(self, event, **kw):
        employee_id = kw.get("employee_id", -1)
        shard_id    = kw.get("shard_id", 0)
        error       = kw.get("error", "unknown")
        run_id      = kw.get("payroll_run_id")

        if not shard_id or shard_id < 0:
            # System-level error; log to all shards
            for sid in self.sm.list_shards():
                self._write(sid, dict(
                    payroll_run_id=run_id, employee_id=employee_id,
                    event_type="ERROR", rule_applied="WORKER_RUNTIME",
                    field_name="error", before_value=None, after_value=error,
                    notes=f"Worker error: {error[:500]}", operator=self.operator,
                ))
            return

        self._write(shard_id, dict(
            payroll_run_id = run_id,
            employee_id    = employee_id,
            event_type     = "ERROR",
            rule_applied   = "WORKER_RUNTIME",
            field_name     = "error",
            before_value   = None,
            after_value    = error[:500],
            notes          = f"Employee {employee_id} failed permanently",
            operator       = self.operator,
        ))

    def on_template_populated(self, event, **kw):
        portal    = kw.get("portal_code", "unknown")
        file_path = kw.get("filled_path", "")
        month     = kw.get("month")
        year      = kw.get("year")
        for sid in self.sm.list_shards():
            self._write(sid, dict(
                payroll_run_id=None, employee_id=None,
                event_type="TEMPLATE", rule_applied=f"PHASE4_{portal.upper()}",
                field_name="filled_path", before_value=None,
                after_value=str(file_path),
                notes=f"Template populated: {portal} for {year}-{month:02d}" if month else f"Template: {portal}",
                operator=self.operator,
            ))

    def on_run_completed(self, event, **kw):
        collector = kw.get("collector")
        month     = kw.get("month")
        year      = kw.get("year")
        elapsed   = kw.get("elapsed_seconds", 0)
        run_map   = kw.get("run_map", {})

        summary = ""
        if collector:
            t = collector.totals
            summary = (
                f"Run {year}-{month:02d} | "
                f"employees={collector.processed_count} | "
                f"errors={len(collector.errors)} | "
                f"gross={t.get('gross_salary',0):.2f} | "
                f"paye={t.get('paye',0):.2f} | "
                f"elapsed={elapsed:.2f}s"
            )

        for shard_id, run_id in run_map.items():
            self._write(shard_id, dict(
                payroll_run_id = run_id,
                employee_id    = None,
                event_type     = "RUN_SUMMARY",
                rule_applied   = "PAYROLL_RUNNER",
                field_name     = "run_status",
                before_value   = "processing",
                after_value    = "completed",
                notes          = summary,
                operator       = self._run_operator.get(str(run_id), self.operator),
            ))

    def on_export_created(self, event, **kw):
        path        = kw.get("path", "")
        export_type = kw.get("export_type", "unknown")
        for sid in self.sm.list_shards():
            self._write(sid, dict(
                payroll_run_id=None, employee_id=None,
                event_type="EXPORT", rule_applied=f"PHASE5_{export_type.upper()}",
                field_name="export_path", before_value=None, after_value=str(path),
                notes=f"Export created: {export_type}", operator=self.operator,
            ))

    # ── Internal ──────────────────────────────────────────────────────────────
    def _write(self, shard_id: int, log: dict) -> None:
        try:
            # Ensure required keys exist with defaults
            log.setdefault("payroll_run_id", None)
            log.setdefault("employee_id", None)
            log.setdefault("rule_applied", None)
            log.setdefault("field_name", None)
            log.setdefault("before_value", None)
            log.setdefault("after_value", None)
            log.setdefault("notes", None)
            log.setdefault("operator", self.operator)
            self.sm.write_audit(shard_id, log)
        except Exception as exc:
            logger.warning(f"AuditPlugin: failed to write to shard {shard_id}: {exc}")


# ─── BACKUP PLUGIN ────────────────────────────────────────────────────────────

class BackupPlugin(PayrollPlugin):
    """
    Snapshots all shard .db files after each aggregation using SQLite's
    native backup API (safe while the database is live and open).

    Layout:
      backups/{YYYY-MM}/shard_{id:04d}_{timestamp}.db
      backups/{YYYY-MM}/manifest_{timestamp}.json    ← SHA-256 checksums

    Retention:
      Purges backup folders older than BACKUP_RETENTION_DAYS (default 90).
      Uses mtime of the folder's manifest file for age calculation.
    """
    name = "BackupPlugin"

    def __init__(self, shard_manager: ShardManager,
                 backup_dir: Path = BACKUP_DIR,
                 retention_days: int = BACKUP_RETENTION_DAYS):
        self.sm            = shard_manager
        self.backup_dir    = Path(backup_dir)
        self.retention_days = retention_days
        self._lock         = threading.Lock()

    def register(self, bus: EventBus | None = None) -> None:
        b = bus or get_bus()
        b.subscribe(PayrollEvent.AGGREGATION_DONE, self.on_aggregation_done)
        logger.info(f"Plugin '{self.name}' registered")

    def on_aggregation_done(self, event, **kw):
        report = kw.get("report")
        month  = kw.get("month")
        year   = kw.get("year")
        if month is None or year is None:
            return
        try:
            with self._lock:
                self._backup(month, year)
                self._purge_old_backups()
        except Exception as exc:
            logger.error(f"BackupPlugin: backup failed: {exc}")

    def backup_now(self, month: int, year: int) -> dict:
        """Public API for manual/scheduled backups. Returns manifest dict."""
        with self._lock:
            result = self._backup(month, year)
            self._purge_old_backups()
        return result

    # ── Internal ──────────────────────────────────────────────────────────────
    def _backup(self, month: int, year: int) -> dict:
        ts       = _now_ts()
        dest_dir = self.backup_dir / _period(month, year)
        dest_dir.mkdir(parents=True, exist_ok=True)

        manifest = {
            "period":     _period(month, year),
            "timestamp":  ts,
            "shards":     [],
        }

        for shard_id in self.sm.list_shards():
            src_path  = self.sm.shard_path(shard_id)
            dest_path = dest_dir / f"shard_{shard_id:04d}_{ts}.db"

            # Use SQLite's native backup API — safe while DB is live
            try:
                src_conn  = sqlite3.connect(str(src_path))
                dest_conn = sqlite3.connect(str(dest_path))
                src_conn.backup(dest_conn)
                dest_conn.close()
                src_conn.close()

                checksum = _sha256(dest_path)
                size_kb  = dest_path.stat().st_size // 1024

                manifest["shards"].append({
                    "shard_id": shard_id,
                    "file":     dest_path.name,
                    "sha256":   checksum,
                    "size_kb":  size_kb,
                })
                logger.info(f"BackupPlugin: shard {shard_id:04d} → {dest_path.name} "
                            f"({size_kb} KB, sha256={checksum[:16]}…)")

            except Exception as exc:
                logger.error(f"BackupPlugin: shard {shard_id} backup failed: {exc}")
                manifest["shards"].append({
                    "shard_id": shard_id,
                    "error":    str(exc),
                })

        # Write manifest
        manifest_path = dest_dir / f"manifest_{ts}.json"
        with open(manifest_path, "w") as f:
            json.dump(manifest, f, indent=2)

        logger.info(f"BackupPlugin: {len(manifest['shards'])} shards backed up "
                    f"to {dest_dir}")
        return manifest

    def _purge_old_backups(self) -> None:
        """Delete period folders whose manifest is older than retention_days."""
        if not self.backup_dir.exists():
            return
        cutoff = datetime.utcnow().timestamp() - (self.retention_days * 86400)
        purged = 0
        for period_dir in self.backup_dir.iterdir():
            if not period_dir.is_dir():
                continue
            manifests = list(period_dir.glob("manifest_*.json"))
            if not manifests:
                continue
            oldest_mtime = min(m.stat().st_mtime for m in manifests)
            if oldest_mtime < cutoff:
                shutil.rmtree(period_dir, ignore_errors=True)
                purged += 1
                logger.info(f"BackupPlugin: purged old backup {period_dir.name}")
        if purged:
            logger.info(f"BackupPlugin: purged {purged} expired backup folder(s)")

    def verify_backup(self, period: str) -> dict:
        """
        Verify SHA-256 checksums of a backup period against its manifest.
        Returns {shard_id: 'ok'|'mismatch'|'missing'}.
        """
        dest_dir = self.backup_dir / period
        results  = {}
        manifests = sorted(dest_dir.glob("manifest_*.json")) if dest_dir.exists() else []
        if not manifests:
            return {"error": f"No manifest found for period {period}"}

        with open(manifests[-1]) as f:
            manifest = json.load(f)

        for entry in manifest.get("shards", []):
            sid      = entry["shard_id"]
            filename = entry.get("file")
            expected = entry.get("sha256")
            if not filename or not expected:
                results[sid] = "missing"
                continue
            backup_file = dest_dir / filename
            if not backup_file.exists():
                results[sid] = "missing"
            elif _sha256(backup_file) == expected:
                results[sid] = "ok"
            else:
                results[sid] = "mismatch"
                logger.error(f"BackupPlugin: TAMPER DETECTED shard {sid} in {period}")
        return results


# ─── P9A ANNUAL AGGREGATOR ────────────────────────────────────────────────────

@dataclass
class P9AEmployeeYear:
    """Annual P9A tax certificate data for one employee."""
    employee_id:           int
    full_name:             str
    kra_pin:               str   = ""
    nssf_number:           str   = ""
    sha_number:            str   = ""
    department:            str   = ""
    job_title:             str   = ""
    tax_year:              int   = 0
    months_worked:         int   = 0

    # Annual totals
    annual_gross:          float = 0.0
    annual_nssf:           float = 0.0
    annual_shif:           float = 0.0
    annual_ahl:            float = 0.0
    annual_taxable_income: float = 0.0
    annual_gross_paye:     float = 0.0
    annual_personal_relief:float = 0.0
    annual_insurance_relief:float = 0.0
    annual_mortgage_relief: float = 0.0
    annual_ahl_relief:      float = 0.0
    annual_post_ret_relief: float = 0.0
    annual_total_relief:    float = 0.0
    annual_paye:            float = 0.0
    annual_net_pay:         float = 0.0

    # Monthly breakdown for P9A form
    monthly_rows:          list  = field(default_factory=list)


class P9AAnnualAggregator:
    """
    Reads all processed monthly payroll runs for a given tax year and
    produces per-employee annual P9A tax certificates.

    Kenya tax year: January–December (calendar year).
    Employers must issue P9A forms by end of February of the following year.

    Usage:
        agg = P9AAnnualAggregator(shard_manager)
        p9a_records = agg.aggregate(year=2025)
        path = agg.export_excel(p9a_records, year=2025)
    """

    MONTH_NAMES = [
        "", "Jan","Feb","Mar","Apr","May","Jun",
        "Jul","Aug","Sep","Oct","Nov","Dec"
    ]

    def __init__(self, shard_manager: ShardManager,
                 export_dir: Path = EXPORT_DIR):
        self.sm         = shard_manager
        self.export_dir = Path(export_dir)

    def aggregate(self, year: int,
                  company_id: int = 1) -> list[P9AEmployeeYear]:
        """
        Aggregate all 12 months of payroll for each employee in the given year.
        Returns list of P9AEmployeeYear, sorted by employee_id.
        """
        records: dict[int, P9AEmployeeYear] = {}

        for shard_id in self.sm.list_shards():
            # Get all processed runs for this year in this shard
            with self.sm.connection(shard_id) as conn:
                run_rows = conn.execute("""
                    SELECT payroll_run_id, run_month FROM payroll_runs
                    WHERE run_year=? AND company_id=?
                      AND status IN ('processed','locked')
                    ORDER BY run_month
                """, (year, company_id)).fetchall()

                if not run_rows:
                    continue

                run_id_map = {r["payroll_run_id"]: r["run_month"] for r in run_rows}
                run_ids    = list(run_id_map.keys())

                # Fetch all transactions for these runs in one query
                placeholders = ",".join("?" * len(run_ids))
                tx_rows = conn.execute(f"""
                    SELECT pt.*, e.full_name, e.kra_pin, e.nssf_number,
                           e.sha_number, e.department, e.job_title
                    FROM payroll_transactions pt
                    JOIN employees e ON e.employee_id = pt.employee_id
                    WHERE pt.payroll_run_id IN ({placeholders})
                    ORDER BY pt.employee_id, pt.payroll_run_id
                """, run_ids).fetchall()

            for row in tx_rows:
                r       = dict(row)
                eid     = r["employee_id"]
                month   = run_id_map.get(r["payroll_run_id"], 0)

                if eid not in records:
                    records[eid] = P9AEmployeeYear(
                        employee_id = eid,
                        full_name   = r.get("full_name", f"Emp {eid}"),
                        kra_pin     = r.get("kra_pin", ""),
                        nssf_number = r.get("nssf_number", ""),
                        sha_number  = r.get("sha_number", ""),
                        department  = r.get("department", ""),
                        job_title   = r.get("job_title", ""),
                        tax_year    = year,
                    )

                rec = records[eid]
                rec.months_worked         += 1
                rec.annual_gross          += r.get("gross_salary", 0)
                rec.annual_nssf           += r.get("nssf_employee", 0)
                rec.annual_shif           += r.get("shif", 0)
                rec.annual_ahl            += r.get("housing_levy_employee", 0)
                rec.annual_taxable_income += r.get("taxable_income", 0)
                rec.annual_gross_paye     += r.get("gross_paye", 0)
                rec.annual_personal_relief+= r.get("personal_relief", 0)
                rec.annual_insurance_relief+=r.get("insurance_relief", 0)
                rec.annual_mortgage_relief += r.get("mortgage_relief", 0)
                rec.annual_ahl_relief     += r.get("ahl_relief", 0)
                rec.annual_post_ret_relief += r.get("post_retirement_relief", 0)
                rec.annual_total_relief   += r.get("total_relief", 0)
                rec.annual_paye           += r.get("paye", 0)
                rec.annual_net_pay        += r.get("net_pay", 0)
                rec.monthly_rows.append({
                    "month": month,
                    "month_name": self.MONTH_NAMES[month] if 1 <= month <= 12 else str(month),
                    "gross":  r.get("gross_salary", 0),
                    "nssf":   r.get("nssf_employee", 0),
                    "shif":   r.get("shif", 0),
                    "ahl":    r.get("housing_levy_employee", 0),
                    "taxable_income": r.get("taxable_income", 0),
                    "gross_paye":  r.get("gross_paye", 0),
                    "total_relief": r.get("total_relief", 0),
                    "paye":   r.get("paye", 0),
                    "net_pay": r.get("net_pay", 0),
                })

        # Round all annual totals to 2dp
        for rec in records.values():
            for f in ("annual_gross","annual_nssf","annual_shif","annual_ahl",
                      "annual_taxable_income","annual_gross_paye","annual_total_relief",
                      "annual_paye","annual_net_pay"):
                setattr(rec, f, round(getattr(rec, f), 2))

        result = sorted(records.values(), key=lambda r: r.employee_id)
        logger.info(f"P9AAnnualAggregator: {len(result)} employees for year {year}")
        return result

    def export_excel(self, records: list[P9AEmployeeYear],
                     year: int,
                     company_info: dict | None = None) -> Path:
        """
        Export P9A forms to Excel — one sheet per employee + an index sheet.
        File: exports/p9a_annual_{year}.xlsx
        """
        self.export_dir.mkdir(parents=True, exist_ok=True)
        company_info = company_info or {}
        wb = openpyxl.Workbook()

        # Index sheet
        ws_idx = wb.active
        ws_idx.title = "P9A Index"
        self._write_index(ws_idx, records, year, company_info)

        # One sheet per employee
        for rec in records:
            ws = wb.create_sheet(title=f"{rec.employee_id:04d}_{rec.full_name[:15]}")
            self._write_p9a_sheet(ws, rec, year, company_info)

        out = self.export_dir / f"p9a_annual_{year}.xlsx"
        wb.save(str(out))
        logger.info(f"P9AAnnualAggregator: saved {out} ({len(records)} certificates)")
        return out

    def _write_index(self, ws, records, year, ci):
        GREEN, DARK = "22c55e", "050e09"
        ws["A1"] = f"P9A Annual Tax Certificates — {ci.get('name','Company')} — Year {year}"
        ws["A1"].font = Font(bold=True, size=13, name="Calibri", color=GREEN)
        ws["A1"].fill = PatternFill("solid", fgColor=DARK)
        ws.merge_cells("A1:K1")
        ws.row_dimensions[1].height = 28

        headers = ["#","Emp ID","Full Name","KRA PIN","NSSF No.","Months",
                   "Annual Gross","Annual PAYE","Annual NSSF","Annual SHIF","Annual Net"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.fill   = PatternFill("solid", fgColor="1a2e20")
            cell.font   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            cell.border = make_border()

        for i, rec in enumerate(records, 1):
            r = i + 3
            vals = [i, rec.employee_id, rec.full_name, rec.kra_pin,
                    rec.nssf_number, rec.months_worked,
                    rec.annual_gross, rec.annual_paye,
                    rec.annual_nssf, rec.annual_shif, rec.annual_net_pay]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill = PatternFill("solid", fgColor="101f15" if i%2 else "0d1a10")
                cell.font = Font(name="Calibri", size=9, color="dff0e6")
                cell.border = make_border()
                if c > 6:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(
                len(str(col[0].value or "")) + 4, 12)

    def _write_p9a_sheet(self, ws, rec: P9AEmployeeYear, year: int, ci: dict):
        GREEN, DARK, GOLD = "22c55e", "050e09", "f59e0b"

        # Title block
        ws["A1"] = "P9A — EMPLOYEE ANNUAL TAX CERTIFICATE"
        ws["A1"].font = Font(bold=True, size=14, name="Calibri", color=GREEN)
        ws["A1"].fill = PatternFill("solid", fgColor=DARK)
        ws.merge_cells("A1:F1")

        info = [
            ("Employer",     ci.get("name", "N/A")),
            ("Employer PIN", ci.get("pin", "N/A")),
            ("Employee",     rec.full_name),
            ("KRA PIN",      rec.kra_pin or "N/A"),
            ("NSSF No.",     rec.nssf_number or "N/A"),
            ("Department",   rec.department or "N/A"),
            ("Tax Year",     str(year)),
            ("Months Worked",str(rec.months_worked)),
        ]
        for r, (lbl, val) in enumerate(info, 3):
            ws.cell(row=r, column=1, value=lbl).font  = Font(bold=True, size=9, name="Calibri", color="5a7a65")
            ws.cell(row=r, column=2, value=val).font  = Font(size=9, name="Calibri", color="dff0e6")

        # Monthly breakdown table
        header_row = 13
        mhdr = ["Month","Gross","NSSF","SHIF","AHL","Taxable","Gross PAYE","Relief","Net PAYE","Net Pay"]
        for c, h in enumerate(mhdr, 1):
            cell = ws.cell(row=header_row, column=c, value=h)
            cell.fill = PatternFill("solid", fgColor="1a2e20")
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
            cell.border = make_border()

        for i, row in enumerate(sorted(rec.monthly_rows, key=lambda x: x["month"]), 1):
            r = header_row + i
            vals = [row["month_name"], row["gross"], row["nssf"], row["shif"],
                    row["ahl"], row["taxable_income"], row["gross_paye"],
                    row["total_relief"], row["paye"], row["net_pay"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=round(v, 2) if isinstance(v, float) else v)
                cell.fill = PatternFill("solid", fgColor="101f15" if i%2 else "0d1a10")
                cell.font = Font(name="Calibri", size=9, color="dff0e6")
                cell.border = make_border()
                if c > 1:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        # Annual totals row
        total_row = header_row + len(rec.monthly_rows) + 1
        totals = ["ANNUAL TOTAL", rec.annual_gross, rec.annual_nssf, rec.annual_shif,
                  rec.annual_ahl, rec.annual_taxable_income, rec.annual_gross_paye,
                  rec.annual_total_relief, rec.annual_paye, rec.annual_net_pay]
        for c, v in enumerate(totals, 1):
            cell = ws.cell(row=total_row, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor="0b1610")
            cell.font = Font(bold=True, color=GOLD, name="Calibri", size=9)
            cell.border = make_border()
            if c > 1:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

        # Auto-width columns
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 14
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["C"].width = 20
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# ─── CONVENIENCE SETUP ───────────────────────────────────────────────────────

def setup_phase6(shard_manager: ShardManager,
                 bus: EventBus | None = None,
                 operator: str = "system",
                 backup_dir: Path = BACKUP_DIR) -> tuple:
    """
    Register both Phase 6 plugins and return them.
    Call this once at system startup before running payroll.

    Returns: (audit_plugin, backup_plugin)
    """
    b = bus or get_bus()
    audit  = AuditPlugin(shard_manager, operator=operator)
    backup = BackupPlugin(shard_manager, backup_dir=backup_dir)
    audit.register(b)
    backup.register(b)
    logger.info("Phase 6: AuditPlugin and BackupPlugin registered")
    return audit, backup
